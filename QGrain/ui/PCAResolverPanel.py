__all__ = ["PCAResolverPanel"]

import typing

import numpy as np
import openpyxl
import qtawesome as qta
from PySide2.QtCore import Qt, QCoreApplication
from PySide2.QtWidgets import (QDialog, QFileDialog, QGridLayout, QLabel,
                               QMessageBox, QPushButton, QSpinBox)
from QGrain import QGRAIN_VERSION
from QGrain.charts.PCAResultChart import PCAResultChart
from QGrain.models.GrainSizeDataset import GrainSizeDataset
from QGrain.ui.LoadDatasetDialog import LoadDatasetDialog
from QGrain.use_excel import column_to_char, prepare_styles
from sklearn.decomposition import PCA


class PCAResolverPanel(QDialog):
    def __init__(self, parent=None):
        flags = Qt.Window | Qt.WindowTitleHint | Qt.CustomizeWindowHint | Qt.WindowCloseButtonHint
        super().__init__(parent=parent, f=flags)
        self.setWindowTitle(self.tr("PCA Resolver"))
        self.init_ui()
        self.load_dataset_dialog = LoadDatasetDialog(parent=self)
        self.load_dataset_dialog.dataset_loaded.connect(self.on_dataset_loaded)
        self.file_dialog = QFileDialog(self)
        self.msg_box = QMessageBox(self)
        self.msg_box.setWindowFlags(Qt.Drawer)
        self.__dataset = None
        self.last_result = None

    def init_ui(self):
        self.main_layout = QGridLayout(self)
        self.chart = PCAResultChart(toolbar=True)
        self.main_layout.addWidget(self.chart, 0, 0, 1, 2)

        self.load_dataset_button = QPushButton(qta.icon("fa5s.database"), self.tr("Load Dataset"))
        self.load_dataset_button.clicked.connect(lambda: self.load_dataset_dialog.show())
        self.main_layout.addWidget(self.load_dataset_button, 1, 0, 1, 2)

        self.n_components_label = QLabel(self.tr("N<sub>components</sub>"))
        self.n_components_input = QSpinBox()
        self.n_components_input.setMinimum(1)
        self.main_layout.addWidget(self.n_components_label, 2, 0)
        self.main_layout.addWidget(self.n_components_input, 2, 1)

        self.perform_button = QPushButton(qta.icon("ei.ok-sign"), self.tr("Perform"))
        self.perform_button.setEnabled(False)
        self.perform_button.clicked.connect(self.on_perform_clicked)
        self.save_button = QPushButton(qta.icon("fa5s.save"), self.tr("Save"))
        self.save_button.setEnabled(False)
        self.save_button.clicked.connect(self.on_save_clicked)
        self.main_layout.addWidget(self.perform_button, 3, 0)
        self.main_layout.addWidget(self.save_button, 3, 1)

    def show_message(self, title: str, message: str):
        self.msg_box.setWindowTitle(title)
        self.msg_box.setText(message)
        self.msg_box.exec_()

    def show_info(self, message: str):
        self.show_message(self.tr("Info"), message)

    def show_warning(self, message: str):
        self.show_message(self.tr("Warning"), message)

    def show_error(self, message: str):
        self.show_message(self.tr("Error"), message)

    def on_dataset_loaded(self, dataset: GrainSizeDataset):
        self.__dataset = dataset
        self.perform_button.setEnabled(True)
        self.n_components_input.setMaximum(len(self.__dataset.classes_μm))

    @property
    def n_components(self):
        return self.n_components_input.value()

    def on_perform_clicked(self):
        X = self.__dataset.X
        pca = PCA(n_components=self.n_components)
        # n_samples x n_components
        transformed = pca.fit_transform(X)
        self.chart.show_result(self.__dataset, transformed, pca)
        self.last_result = (self.__dataset, transformed, pca.components_)
        self.save_button.setEnabled(True)

    def on_save_clicked(self):
        if self.last_result is None:
            self.show_warning(self.tr("The PCA algorithm has not been performed."))
            return
        filename, _ = self.file_dialog.getSaveFileName(
            None, self.tr("Save PCA Result"),
            None, "Microsoft Excel (*.xlsx)")
        if filename is None or filename == "":
            return
        try:
            self.save_as_xlsx(filename)
            self.show_info(self.tr("PCA result has been saved to:\n    {0}").format(filename))
        except Exception as e:
            self.show_error(self.tr("Error raised while save PCA result to Excel file.\n    {0}").format(e.__str__()))

    def save_as_xlsx(self, filename: str):
        assert self.last_result is not None
        dataset, transformed, components = self.last_result
        n_samples, n_components = transformed.shape

        wb = openpyxl.Workbook()
        prepare_styles(wb)
        ws = wb.active
        ws.title = self.tr("README")
        description = \
            """
            This Excel file was generated by QGrain ({0}).

            It contanins three sheets:
            1. The first sheet is the dataset which was used to perform the PCA algorithm.
            2. The second sheet is used to put the distributions of all PCs.
            3. The third sheet is the PC variation of all samples.

            The base PCA algorithm is implemented by scikit-learn. You can get the details of algorithm from the following website.
            https://scikit-learn.org/stable/modules/generated/sklearn.decomposition.PCA.html

            """.format(QGRAIN_VERSION)

        def write(row, col, value, style="normal_light"):
            cell = ws.cell(row+1, col+1, value=value)
            cell.style = style

        lines_of_desc = description.split("\n")
        for row, line in enumerate(lines_of_desc):
            write(row, 0, line, style="description")
        ws.column_dimensions[column_to_char(0)].width = 200

        ws = wb.create_sheet(self.tr("Dataset"))
        write(0, 0, self.tr("Sample Name"), style="header")
        ws.column_dimensions[column_to_char(0)].width = 16
        for col, value in enumerate(dataset.classes_μm, 1):
            write(0, col, value, style="header")
            ws.column_dimensions[column_to_char(col)].width = 10
        for row, sample in enumerate(dataset.samples, 1):
            if row % 2 == 0:
                style = "normal_dark"
            else:
                style = "normal_light"
            write(row, 0, sample.name, style=style)
            for col, value in enumerate(sample.distribution, 1):
                write(row, col, value, style=style)
            QCoreApplication.processEvents()

        ws = wb.create_sheet(self.tr("PCs"))
        write(0, 0, self.tr("PC"), style="header")
        ws.column_dimensions[column_to_char(0)].width = 16
        for col, value in enumerate(dataset.classes_μm, 1):
            write(0, col, value, style="header")
            ws.column_dimensions[column_to_char(col)].width = 10
        for row, component in enumerate(components, 1):
            if row % 2 == 0:
                style = "normal_dark"
            else:
                style = "normal_light"
            write(row, 0, self.tr("PC{0}").format(row), style=style)
            for col, value in enumerate(component, 1):
                write(row, col, value, style=style)
            QCoreApplication.processEvents()

        ws = wb.create_sheet(self.tr("Variations of PCs"))
        write(0, 0, self.tr("Sample Name"), style="header")
        ws.column_dimensions[column_to_char(0)].width = 16
        for i in range(n_components):
            write(0, i+1, self.tr("PC{0}").format(i+1), style="header")
            ws.column_dimensions[column_to_char(i+1)].width = 10
        for row, varations in enumerate(transformed, 1):
            if row % 2 == 0:
                style = "normal_dark"
            else:
                style = "normal_light"
            write(row, 0, dataset.samples[row-1].name, style=style)
            for col, value in enumerate(varations, 1):
                write(row, col, value, style=style)
            QCoreApplication.processEvents()

        wb.save(filename)
        wb.close()


if __name__ == "__main__":
    import sys

    from QGrain.entry import setup_app
    app = setup_app()
    main = PCAResolverPanel()
    main.show()
    sys.exit(app.exec_())