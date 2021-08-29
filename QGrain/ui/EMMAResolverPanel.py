__all__ = ["EMMAResolverPanel"]

import os
import pickle
import typing

import numpy as np
import openpyxl
import qtawesome as qta
from PySide2.QtCore import QCoreApplication, Qt
from PySide2.QtWidgets import (QComboBox, QDialog, QFileDialog, QGridLayout,
                               QGroupBox, QLabel, QListWidget, QMessageBox,
                               QProgressBar, QPushButton, QSpinBox)
from QGrain import QGRAIN_VERSION
from QGrain import DistributionType
from QGrain.emma import EMMAResolver
from QGrain.statistic import convert_μm_to_φ
from QGrain.charts.EMMAResultChart import EMMAResultChart
from QGrain.charts.EMMASummaryChart import EMMASummaryChart
from QGrain.emma import EMMAResult
from QGrain.models.GrainSizeDataset import GrainSizeDataset
from QGrain.models.NNResolverSetting import NNResolverSetting
from QGrain.ui.LoadDatasetDialog import LoadDatasetDialog
from QGrain.ui.NNResolverSettingWidget import NNResolverSettingWidget
from QGrain.use_excel import column_to_char, prepare_styles


class EMMAResolverPanel(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent=parent, f=Qt.Window)
        self.setWindowTitle(self.tr("EMMA Resolver"))
        self.SUPPORTED_DISTS = \
            [(DistributionType.Nonparametric, self.tr("Nonparametric")),
             (DistributionType.Normal, self.tr("Normal")),
             (DistributionType.Weibull, self.tr("Weibull")),
             (DistributionType.SkewNormal, self.tr("Skew Normal"))]

        self.init_ui()
        self.normal_msg = QMessageBox(self)
        self.__dataset = None # type: GrainSizeDataset
        self.__result_list = [] # type: list[EMMAResult]
        self.neural_setting = NNResolverSettingWidget(parent=self)
        self.neural_setting.setting = NNResolverSetting(min_niter=800, max_niter=1200, tol=1e-6, ftol=1e-8, lr=5e-2)
        self.load_dialog = LoadDatasetDialog(parent=self)
        self.load_dialog.dataset_loaded.connect(self.on_dataset_loaded)
        self.file_dialog = QFileDialog(parent=self)
        self.emma_result_chart = EMMAResultChart(toolbar=True)
        self.emma_summary_chart = EMMASummaryChart(toolbar=True)

    def init_ui(self):
        self.setAttribute(Qt.WA_StyledBackground, True)
        self.main_layout = QGridLayout(self)
        # self.main_layout.setContentsMargins(0, 0, 0, 0)
        # control group
        self.control_group = QGroupBox(self.tr("Control"))
        self.control_layout = QGridLayout(self.control_group)
        self.main_layout.addWidget(self.control_group, 0, 0)
        self.load_dataset_button = QPushButton(qta.icon("fa.database"), self.tr("Load Dataset"))
        self.load_dataset_button.clicked.connect(self.on_load_clicked)
        self.control_layout.addWidget(self.load_dataset_button, 0, 0, 1, 2)
        self.configure_button = QPushButton(qta.icon("fa.gears"), self.tr("Configure Algorithm"))
        self.configure_button.clicked.connect(self.on_configure_clicked)
        self.control_layout.addWidget(self.configure_button, 1, 0, 1, 2)
        self.n_samples_label = QLabel(self.tr("N<sub>samples</sub>"))
        self.n_samples_display = QLabel(self.tr("Unknown"))
        self.control_layout.addWidget(self.n_samples_label, 2, 0)
        self.control_layout.addWidget(self.n_samples_display, 2, 1)
        self.distribution_label = QLabel(self.tr("Distribution Type"))
        self.distribution_combo_box = QComboBox()
        self.distribution_combo_box.addItems([name for _, name in self.SUPPORTED_DISTS])
        self.control_layout.addWidget(self.distribution_label, 3, 0)
        self.control_layout.addWidget(self.distribution_combo_box, 3, 1)
        self.min_n_members_label = QLabel("Minimum N<sub>members</sub>")
        self.min_n_members_input = QSpinBox()
        self.min_n_members_input.setRange(1, 10)
        self.control_layout.addWidget(self.min_n_members_label, 4, 0)
        self.control_layout.addWidget(self.min_n_members_input, 4, 1)
        self.max_n_members_label = QLabel("Maximum N<sub>members</sub>")
        self.max_n_members_input = QSpinBox()
        self.max_n_members_input.setRange(1, 10)
        self.max_n_members_input.setValue(10)
        self.control_layout.addWidget(self.max_n_members_label, 5, 0)
        self.control_layout.addWidget(self.max_n_members_input, 5, 1)
        self.perform_button = QPushButton(qta.icon("fa.play-circle"), self.tr("Perform"))
        self.perform_button.clicked.connect(self.on_perform_clicked)
        self.perform_button.setEnabled(False)
        self.perform_with_customized_ems_button = QPushButton(qta.icon("fa.play-circle"), self.tr("Perform With Customized EMs"))
        self.perform_with_customized_ems_button.clicked.connect(self.on_perform_with_customized_ems)
        self.perform_with_customized_ems_button.setEnabled(False)
        self.control_layout.addWidget(self.perform_button, 6, 0, 1, 2)
        self.control_layout.addWidget(self.perform_with_customized_ems_button, 7, 0, 1, 2)
        self.progress_bar = QProgressBar()
        self.progress_bar.setFormat(self.tr("EMMA Progress"))
        self.control_layout.addWidget(self.progress_bar, 8, 0, 1, 2)

        self.result_group = QGroupBox(self.tr("Result"))
        self.result_layout = QGridLayout(self.result_group)
        self.main_layout.addWidget(self.result_group, 0, 1)
        self.result_list_widget = QListWidget()
        self.result_layout.addWidget(self.result_list_widget, 0, 0, 1, 2)
        self.remove_result_button = QPushButton(qta.icon("mdi.delete"), self.tr("Remove"))
        self.remove_result_button.clicked.connect(self.on_remove_clicked)
        self.show_result_button = QPushButton(qta.icon("fa.area-chart"), self.tr("Show"))
        self.show_result_button.clicked.connect(self.on_show_clicked)
        self.load_dump_button = QPushButton(qta.icon("fa.database"), self.tr("Load Dump"))
        self.load_dump_button.clicked.connect(self.on_load_dump_clicked)
        self.save_button = QPushButton(qta.icon("fa.save"), self.tr("Save"))
        self.save_button.clicked.connect(self.on_save_clicked)
        self.remove_result_button.setEnabled(False)
        self.show_result_button.setEnabled(False)
        self.save_button.setEnabled(False)
        self.result_layout.addWidget(self.remove_result_button, 1, 0)
        self.result_layout.addWidget(self.show_result_button, 1, 1)
        self.result_layout.addWidget(self.load_dump_button, 2, 0)
        self.result_layout.addWidget(self.save_button, 2, 1)

    def show_message(self, title: str, message: str):
        self.normal_msg.setWindowTitle(title)
        self.normal_msg.setText(message)
        self.normal_msg.exec_()

    def show_info(self, message: str):
        self.show_message(self.tr("Info"), message)

    def show_warning(self, message: str):
        self.show_message(self.tr("Warning"), message)

    def show_error(self, message: str):
        self.show_message(self.tr("Error"), message)

    def on_dataset_loaded(self, dataset: GrainSizeDataset):
        self.__dataset = dataset
        self.n_samples_display.setText(str(self.__dataset.n_samples))
        self.perform_button.setEnabled(True)
        self.perform_with_customized_ems_button.setEnabled(True)

    def on_load_clicked(self):
        self.load_dialog.show()

    def on_configure_clicked(self):
        self.neural_setting.show()

    @property
    def distribution_type(self) -> DistributionType:
        distribution_type, _ = self.SUPPORTED_DISTS[self.distribution_combo_box.currentIndex()]
        return distribution_type

    @property
    def n_members_list(self):
        min_n = self.min_n_members_input.value()
        max_n = self.max_n_members_input.value()
        if min_n > max_n:
            min_n, max_n = max_n, min_n
        return list(range(min_n, max_n+1, 1))

    @property
    def n_results(self) -> int:
        return len(self.__result_list)

    @property
    def selected_index(self):
        indexes = self.result_list_widget.selectedIndexes()
        if len(indexes) == 0:
            return 0
        else:
            return indexes[0].row()

    @property
    def selected_result(self):
        if self.n_results == 0:
            None
        else:
            return self.__result_list[self.selected_index]

    def on_perform_clicked(self):
        if self.__dataset is None:
            self.show_error(self.tr("Dataset has not been loaded."))
            return

        self.perform_button.setEnabled(False)
        self.perform_with_customized_ems_button.setEnabled(False)
        resolver = EMMAResolver()
        resolver_setting = self.neural_setting.setting
        results = []
        n_members_list = self.n_members_list
        self.progress_bar.setRange(0, len(n_members_list))
        self.progress_bar.setValue(0)
        self.progress_bar.setFormat(self.tr("Performing EMMA [%v/%m]"))
        QCoreApplication.processEvents()
        for i, n_members in enumerate(n_members_list):
            result = resolver.try_fit(self.__dataset, self.distribution_type, n_members, resolver_setting)
            results.append(result)
            self.progress_bar.setValue(i+1)
            QCoreApplication.processEvents()

        self.add_results(results)
        self.progress_bar.setFormat(self.tr("Finished"))
        self.perform_button.setEnabled(True)
        self.perform_with_customized_ems_button.setEnabled(True)
        if len(results) > 1:
            self.emma_summary_chart.show_distances(results)
            self.emma_summary_chart.show()

    def on_perform_with_customized_ems(self):
        if self.__dataset is None:
            self.show_error(self.tr("Dataset has not been loaded."))
            return

        filename, _ = self.file_dialog.getOpenFileName(
            self, self.tr("Choose a excel file which contains the customized EMs at the first sheet"),
            None, f"{self.tr('Microsoft Excel')} (*.xlsx)")
        if filename is None or filename == "":
            return
        try:
            wb = openpyxl.load_workbook(filename, read_only=True, data_only=True)
            ws = wb[wb.sheetnames[0]]
            raw_data = [[value for value in row] for row in ws.values]
            classes_μm = np.array(raw_data[0][1:], dtype=np.float64)
            classes_φ = convert_μm_to_φ(classes_μm)
            em_distributions = [np.array(row[1:], dtype=np.float64) for row in raw_data[1:]]
        except Exception as e:
            self.show_error(self.tr("Error raised while loading the customized EMs.\n    {0}").format(e.__str__()))
            return
        if len(classes_μm) < 10:
            self.show_error(self.tr("The length of grain-size classes is too less."))
            return
        for i in range(len(classes_μm)-1):
            if classes_μm[i+1] <= classes_μm[i]:
                self.show_error(self.tr("The grain-size classes is not incremental."))
                return
        if np.any(np.isnan(classes_μm)):
            self.show_error(self.tr("There is at least one nan value in grain-size classes."))
            return
        if len(em_distributions) > 10:
            self.show_error(self.tr("There are more than 10 customized EMs in the first sheet, please check."))
            return
        for distribution in em_distributions:
            if len(classes_μm) != len(distribution):
                self.show_error(self.tr("Some distributions of customized EMs have different length with the grain-size classes."))
                return
            if np.any(np.isnan(distribution)):
                self.show_error(self.tr("There is at least one nan value in the frequceny distributions of EMs."))
                return
            if abs(np.sum(distribution) - 1.0) > 0.05:
                self.show_error(self.tr("The sum of some distributions of customized EMs are not equal to 1."))
                return

        self.perform_button.setEnabled(False)
        self.perform_with_customized_ems_button.setEnabled(False)
        resolver = EMMAResolver()
        resolver_setting = self.neural_setting.setting

        self.progress_bar.setRange(0, 1)
        self.progress_bar.setValue(0)
        self.progress_bar.setFormat(self.tr("Performing EMMA [%v/%m]"))
        QCoreApplication.processEvents()
        try:
            result = resolver.try_fit_with_fixed_ems(self.__dataset, classes_φ, em_distributions, resolver_setting)
            self.progress_bar.setValue(1)
            self.add_results([result])
            self.progress_bar.setFormat(self.tr("Finished"))
        except Exception as e:
            self.show_error(self.tr("Error raised while fitting.\n    {0}").format(e.__str__()))
            self.progress_bar.setFormat(self.tr("Failed"))
        QCoreApplication.processEvents()
        self.perform_button.setEnabled(True)
        self.perform_with_customized_ems_button.setEnabled(True)

    def get_distribution_name(self, distribution_type: DistributionType):
        if distribution_type == DistributionType.Nonparametric:
            return self.tr("Nonparametric")
        elif distribution_type == DistributionType.Normal:
            return self.tr("Normal")
        elif distribution_type == DistributionType.Weibull:
            return self.tr("Weibull")
        elif distribution_type == DistributionType.SkewNormal:
            return self.tr("Skew Normal")
        elif distribution_type == DistributionType.Customized:
            return self.tr("Customized")
        else:
            raise NotImplementedError(distribution_type)

    def get_result_name(self, result: EMMAResult):
        return f"[{self.get_distribution_name(result.distribution_type)}]-[{result.n_members} EM(s)]"

    def add_results(self, results: typing.List[EMMAResult]):
        if self.n_results == 0:
            self.remove_result_button.setEnabled(True)
            self.show_result_button.setEnabled(True)
            self.save_button.setEnabled(True)

        self.__result_list.extend(results)
        self.result_list_widget.addItems([self.get_result_name(result) for result in results])

    def on_remove_clicked(self):
        if self.n_results == 0:
            return
        else:
            index = self.selected_index
            self.__result_list.pop(index)
            self.result_list_widget.takeItem(index)

        if self.n_results == 0:
            self.remove_result_button.setEnabled(False)
            self.show_result_button.setEnabled(False)
            self.save_button.setEnabled(False)

    def on_show_clicked(self):
        result = self.selected_result
        if result is None:
            return
        else:
            self.emma_result_chart.show_result(result)
            self.emma_result_chart.show()

    def on_load_dump_clicked(self):
        filename, _  = self.file_dialog.getOpenFileName(self, self.tr("Select the dump file of the EMMA result(s)"),
                                         None, f"{self.tr('Binary Dump')} (*.dump)")
        if filename is None or filename == "":
            return
        with open(filename, "rb") as f:
            results = pickle.load(f)
            invalid = False
            if isinstance(results, list):
                for result in results:
                    if not isinstance(result, EMMAResult):
                        invalid = True
                        break
            else:
                invalid = True
            if invalid:
                self.show_error(self.tr("The dump file does not contain any EMMA result."))
                return
            else:
                self.add_results(results)

    def save_result_excel(self, filename: str, result: EMMAResult):
        # get the mode size of each end-members
        modes = [(i, result.dataset.classes_μm[np.unravel_index(np.argmax(result.end_members[i]), result.end_members[i].shape)]) for i in range(result.n_members)]
        # sort them by mode size
        modes.sort(key=lambda x: x[1])
        wb = openpyxl.Workbook()
        prepare_styles(wb)

        ws = wb.active
        ws.title = self.tr("README")
        description = \
            """
            This Excel file was generated by QGrain ({0}).

            Please cite:
            Liu, Y., Liu, X., Sun, Y., 2021. QGrain: An open-source and easy-to-use software for the comprehensive analysis of grain size distributions. Sedimentary Geology 423, 105980. https://doi.org/10.1016/j.sedgeo.2021.105980

            It contanins three sheets:
            1. The first sheet is the dataset which was used to perform the EMMA algorithm.
            2. The second sheet is used to put the distributions of all end-members.
            3. The third sheet is the end-member fractions of all samples.

            This EMMA algorithm was implemented by QGrian, using the famous machine learning framework, PyTorch.

            EMMA algorithm details
                N_samples: {1},
                Distribution Type: {2},
                N_members: {3},
                N_iterations: {4},
                Spent Time: {5} s,

                Computing Device: {6},
                Distance: {7},
                Minimum N_iterations: {8},
                Maximum N_iterations: {9},
                Learning Rate: {10},
                eps: {11},
                tol: {12},
                ftol: {13}

            """.format(QGRAIN_VERSION,
                    result.dataset.n_samples,
                    result.distribution_type.name,
                    result.n_members,
                    result.n_iterations,
                    result.time_spent,
                    result.resolver_setting.device,
                    result.resolver_setting.distance,
                    result.resolver_setting.min_niter,
                    result.resolver_setting.max_niter,
                    result.resolver_setting.lr,
                    result.resolver_setting.eps,
                    result.resolver_setting.tol,
                    result.resolver_setting.ftol)

        def write(row, col, value, style="normal_light"):
            cell = ws.cell(row+1, col+1, value=value)
            cell.style = style

        lines_of_desc = description.split("\n")
        for row, line in enumerate(lines_of_desc):
            write(row, 0, line, style="description")
        ws.column_dimensions[column_to_char(0)].width = 200

        ws = wb.create_sheet(self.tr("Dataset"))
        write(0, 0, self.tr("Sample Name"), style="header")
        ws.column_dimensions[column_to_char(0)].width = 16
        for col, value in enumerate(result.dataset.classes_μm, 1):
            write(0, col, value, style="header")
            ws.column_dimensions[column_to_char(col)].width = 10
        for row, sample in enumerate(result.dataset.samples, 1):
            if row % 2 == 0:
                style = "normal_dark"
            else:
                style = "normal_light"
            write(row, 0, sample.name, style=style)
            for col, value in enumerate(sample.distribution, 1):
                write(row, col, value, style=style)
            QCoreApplication.processEvents()

        ws = wb.create_sheet(self.tr("End-members"))
        write(0, 0, self.tr("End-member"), style="header")
        ws.column_dimensions[column_to_char(0)].width = 16
        for col, value in enumerate(result.dataset.classes_μm, 1):
            write(0, col, value, style="header")
            ws.column_dimensions[column_to_char(col)].width = 10
        for row, (index, _) in enumerate(modes, 1):
            if row % 2 == 0:
                style = "normal_dark"
            else:
                style = "normal_light"
            write(row, 0, f"EM{row}", style=style)
            for col, value in enumerate(result.end_members[index], 1):
                write(row, col, value, style=style)
            QCoreApplication.processEvents()

        ws = wb.create_sheet(self.tr("Fractions"))
        write(0, 0, self.tr("Sample Name"), style="header")
        ws.column_dimensions[column_to_char(0)].width = 16
        for i in range(result.n_members):
            write(0, i+1, f"EM{i+1}", style="header")
            ws.column_dimensions[column_to_char(i+1)].width = 10
        for row, fractions in enumerate(result.fractions, 1):
            if row % 2 == 0:
                style = "normal_dark"
            else:
                style = "normal_light"
            write(row, 0, result.dataset.samples[row-1].name, style=style)
            for col, (index, _) in enumerate(modes, 1):
                write(row, col, fractions[index], style=style)
            QCoreApplication.processEvents()

        wb.save(filename)
        wb.close()

    def on_save_clicked(self):
        if self.n_results == 0:
            self.show_warning(self.tr("There is not an EMMA result in the list."))
            return

        filename, _ = self.file_dialog.getSaveFileName(
            self, self.tr("Choose a filename to save the EMMA result(s) in list"),
            None, f"{self.tr('Binary Dump')} (*.dump);;{self.tr('Microsoft Excel')} (*.xlsx)")
        if filename is None or filename == "":
            return
        _, ext = os.path.splitext(filename)

        if ext == ".dump":
            with open(filename, "wb") as f:
                pickle.dump(self.__result_list, f)
                self.show_info(self.tr("All results in list has been saved."))
        elif ext == ".xlsx":
            try:
                result = self.selected_result
                self.save_result_excel(filename, result)
                self.show_info(self.tr("The selected result has been saved."))
            except Exception as e:
                self.show_error(self.tr("Error raised while saving it to Excel file.\n    {0}").format(e.__str__()))
                return


if __name__ == "__main__":
    import sys

    from QGrain.entry import setup_app
    app, splash = setup_app()
    main = EMMAResolverPanel()
    main.show()
    splash.finish(main)
    sys.exit(app.exec_())
