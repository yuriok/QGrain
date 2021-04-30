__all__ = ["GrainSizeDatasetViewer"]

import logging

import openpyxl
import qtawesome as qta
from PySide2.QtCore import Qt
from PySide2.QtGui import QCursor
from PySide2.QtWidgets import (QAbstractItemView, QCheckBox, QComboBox,
                               QDialog, QFileDialog, QGridLayout, QHeaderView, QMenu,
                               QMessageBox, QPushButton, QTableWidget,
                               QTableWidgetItem)
from QGrain import QGRAIN_VERSION
from QGrain.algorithms.moments import get_moments
from QGrain.charts.CumulativeCurveChart import CumulativeCurveChart
from QGrain.charts.diagrams import (BP12GSMDiagramChart,
                                    BP12SSCDiagramChart,
                                    Folk54GSMDiagramChart,
                                    Folk54SSCDiagramChart)
from QGrain.charts.FrequencyCurve3DChart import FrequencyCurve3DChart
from QGrain.charts.FrequencyCurveChart import FrequencyCurveChart
from QGrain.models.GrainSizeDataset import GrainSizeDataset
from QGrain.ui.LoadDatasetDialog import LoadDatasetDialog
from QGrain.use_excel import column_to_char, prepare_styles


class GrainSizeDatasetViewer(QDialog):
    PAGE_ROWS = 20
    logger = logging.getLogger("root.ui.GrainSizeDatasetView")
    gui_logger = logging.getLogger("GUI")
    def __init__(self, parent=None):
        super().__init__(parent=parent, f=Qt.Window)
        self.setWindowTitle(self.tr("Grain-size Dataset Viewer"))
        self.__dataset = GrainSizeDataset() # type: GrainSizeDataset
        self.init_ui()
        self.data_table.setRowCount(0)
        self.frequency_curve_chart = FrequencyCurveChart(parent=self, toolbar=True)
        self.frequency_curve_3D_chart = FrequencyCurve3DChart(parent=self, toolbar=True)
        self.cumulative_curve_chart = CumulativeCurveChart(parent=self, toolbar=True)
        self.folk54_GSM_diagram_chart = Folk54GSMDiagramChart(parent=self, toolbar=True)
        self.folk54_SSC_diagram_chart = Folk54SSCDiagramChart(parent=self, toolbar=True)
        self.BP12_GSM_diagram_chart = BP12GSMDiagramChart(parent=self, toolbar=True)
        self.BP12_SSC_diagram_chart = BP12SSCDiagramChart(parent=self, toolbar=True)
        self.load_dataset_dialog = LoadDatasetDialog(parent=self)
        self.load_dataset_dialog.dataset_loaded.connect(self.on_data_loaded)
        self.file_dialog = QFileDialog(parent=self)
        self.msg_box = QMessageBox(self)
        self.msg_box.setWindowFlags(Qt.Drawer)

    def init_ui(self):
        self.setWindowTitle(self.tr("Dataset Viewer"))
        self.data_table = QTableWidget(100, 100)
        self.data_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.data_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.data_table.setAlternatingRowColors(True)
        self.data_table.setContextMenuPolicy(Qt.CustomContextMenu)
        # self.data_table.hideColumn(0)
        self.main_layout = QGridLayout(self)
        self.main_layout.addWidget(self.data_table, 0, 0, 1, 3)

        self.previous_button = QPushButton(self.tr("Previous"))
        self.previous_button.setToolTip(self.tr("Click to back to the previous page."))
        self.previous_button.clicked.connect(self.on_previous_button_clicked)
        self.current_page_combo_box = QComboBox()
        self.current_page_combo_box.currentIndexChanged.connect(self.update_page)
        self.next_button = QPushButton(self.tr("Next"))
        self.next_button.setToolTip(self.tr("Click to jump to the next page."))
        self.next_button.clicked.connect(self.on_next_button_clicked)
        self.main_layout.addWidget(self.previous_button, 1, 0)
        self.main_layout.addWidget(self.current_page_combo_box, 1, 1)
        self.main_layout.addWidget(self.next_button, 1, 2)

        self.geometric_checkbox = QCheckBox(self.tr("Geometric"))
        self.geometric_checkbox.setChecked(True)
        self.geometric_checkbox.stateChanged.connect(self.on_is_geometric_changed)
        self.main_layout.addWidget(self.geometric_checkbox, 2, 0)
        self.FW57_checkbox = QCheckBox(self.tr("Folk and Ward (1957) method"))
        self.FW57_checkbox.setChecked(False)
        self.FW57_checkbox.stateChanged.connect(self.on_use_FW57_changed)
        self.main_layout.addWidget(self.FW57_checkbox, 2, 1)
        self.proportion_combo_box = QComboBox()
        self.supported_proportions = [("GSM_proportion", self.tr("Gravel, Sand, Mud")),
                                      ("SSC_proportion", self.tr("Sand, Silt, Clay")),
                                      ("BGSSC_proportion", self.tr("Boulder, Gravel, Sand, Silt, Clay"))]
        self.proportion_combo_box.addItems([description for _, description in self.supported_proportions])
        self.proportion_combo_box.currentIndexChanged.connect(lambda: self.update_page(self.page_index))
        self.main_layout.addWidget(self.proportion_combo_box, 2, 2)

        self.menu = QMenu(self.data_table)
        self.load_dataset_action = self.menu.addAction(qta.icon("fa.database"), self.tr("Load Dataset"))
        self.load_dataset_action.triggered.connect(self.load_dataset)
        self.plot_cumulative_curve_menu = self.menu.addMenu(qta.icon("mdi.chart-bell-curve-cumulative"), self.tr("Plot Cumlulative Curve Chart"))
        self.cumulative_plot_selected_action = self.plot_cumulative_curve_menu.addAction(self.tr("Plot Selected Samples"))
        self.cumulative_plot_selected_action.triggered.connect(lambda: self.plot_chart(self.cumulative_curve_chart, self.selections, False))
        self.cumulative_append_selected_action = self.plot_cumulative_curve_menu.addAction(self.tr("Append Selected Samples"))
        self.cumulative_append_selected_action.triggered.connect(lambda: self.plot_chart(self.cumulative_curve_chart, self.selections, True))
        self.cumulative_plot_all_action = self.plot_cumulative_curve_menu.addAction(self.tr("Plot All Samples"))
        self.cumulative_plot_all_action.triggered.connect(lambda: self.plot_chart(self.cumulative_curve_chart, self.__dataset.samples, False))

        self.plot_frequency_curve_menu = self.menu.addMenu(qta.icon("mdi.chart-bell-curve"), self.tr("Plot Frequency Curve Chart"))
        self.frequency_plot_selected_action = self.plot_frequency_curve_menu.addAction(self.tr("Plot Selected Samples"))
        self.frequency_plot_selected_action.triggered.connect(lambda: self.plot_chart(self.frequency_curve_chart, self.selections, False))
        self.frequency_append_selected_action = self.plot_frequency_curve_menu.addAction(self.tr("Append Selected Samples"))
        self.frequency_append_selected_action.triggered.connect(lambda: self.plot_chart(self.frequency_curve_chart, self.selections, True))
        self.frequency_plot_all_action = self.plot_frequency_curve_menu.addAction(self.tr("Plot All Samples"))
        self.frequency_plot_all_action.triggered.connect(lambda: self.plot_chart(self.frequency_curve_chart, self.__dataset.samples, False))

        self.plot_frequency_curve_3D_menu = self.menu.addMenu(qta.icon("mdi.video-3d"), self.tr("Plot Frequency Curve 3D Chart"))
        self.frequency_3D_plot_selected_action = self.plot_frequency_curve_3D_menu.addAction(self.tr("Plot Selected Samples"))
        self.frequency_3D_plot_selected_action.triggered.connect(lambda: self.plot_chart(self.frequency_curve_3D_chart, self.selections, False))
        self.frequency_3D_append_selected_action = self.plot_frequency_curve_3D_menu.addAction(self.tr("Append Selected Samples"))
        self.frequency_3D_append_selected_action.triggered.connect(lambda: self.plot_chart(self.frequency_curve_3D_chart, self.selections, True))
        self.frequency_3D_plot_all_action = self.plot_frequency_curve_3D_menu.addAction(self.tr("Plot All Samples"))
        self.frequency_3D_plot_all_action.triggered.connect(lambda: self.plot_chart(self.frequency_curve_3D_chart, self.__dataset.samples, False))

        self.folk54_GSM_diagram_menu = self.menu.addMenu(qta.icon("mdi.triangle-outline"), self.tr("Plot GSM Diagram (Folk, 1954)"))
        self.folk54_GSM_plot_selected_action = self.folk54_GSM_diagram_menu.addAction(self.tr("Plot Selected Samples"))
        self.folk54_GSM_plot_selected_action.triggered.connect(lambda: self.plot_chart(self.folk54_GSM_diagram_chart, self.selections, False))
        self.folk54_GSM_append_selected_action = self.folk54_GSM_diagram_menu.addAction(self.tr("Append Selected Samples"))
        self.folk54_GSM_append_selected_action.triggered.connect(lambda: self.plot_chart(self.folk54_GSM_diagram_chart, self.selections, True))
        self.folk54_GSM_plot_all_action = self.folk54_GSM_diagram_menu.addAction(self.tr("Plot All Samples"))
        self.folk54_GSM_plot_all_action.triggered.connect(lambda: self.plot_chart(self.folk54_GSM_diagram_chart, self.__dataset.samples, False))

        self.folk54_SSC_diagram_menu = self.menu.addMenu(qta.icon("mdi.triangle-outline"), self.tr("Plot SSC Diagram (Folk, 1954)"))
        self.folk54_SSC_plot_selected_action = self.folk54_SSC_diagram_menu.addAction(self.tr("Plot Selected Samples"))
        self.folk54_SSC_plot_selected_action.triggered.connect(lambda: self.plot_chart(self.folk54_SSC_diagram_chart, self.selections, False))
        self.folk54_SSC_append_selected_action = self.folk54_SSC_diagram_menu.addAction(self.tr("Append Selected Samples"))
        self.folk54_SSC_append_selected_action.triggered.connect(lambda: self.plot_chart(self.folk54_SSC_diagram_chart, self.selections, True))
        self.folk54_SSC_plot_all_action = self.folk54_SSC_diagram_menu.addAction(self.tr("Plot All Samples"))
        self.folk54_SSC_plot_all_action.triggered.connect(lambda: self.plot_chart(self.folk54_SSC_diagram_chart, self.__dataset.samples, False))

        self.BP12_GSM_diagram_menu = self.menu.addMenu(qta.icon("mdi.triangle-outline"), self.tr("Plot GSM Diagram (Blott && Pye, 2012)"))
        self.BP12_GSM_plot_selected_action = self.BP12_GSM_diagram_menu.addAction(self.tr("Plot Selected Samples"))
        self.BP12_GSM_plot_selected_action.triggered.connect(lambda: self.plot_chart(self.BP12_GSM_diagram_chart, self.selections, False))
        self.BP12_GSM_append_selected_action = self.BP12_GSM_diagram_menu.addAction(self.tr("Append Selected Samples"))
        self.BP12_GSM_append_selected_action.triggered.connect(lambda: self.plot_chart(self.BP12_GSM_diagram_chart, self.selections, True))
        self.BP12_GSM_plot_all_action = self.BP12_GSM_diagram_menu.addAction(self.tr("Plot All Samples"))
        self.BP12_GSM_plot_all_action.triggered.connect(lambda: self.plot_chart(self.BP12_GSM_diagram_chart, self.__dataset.samples, False))

        self.BP12_SSC_diagram_menu = self.menu.addMenu(qta.icon("mdi.triangle-outline"), self.tr("Plot SSC Diagram (Blott && Pye, 2012)"))
        self.BP12_SSC_plot_selected_action = self.BP12_SSC_diagram_menu.addAction(self.tr("Plot Selected Samples"))
        self.BP12_SSC_plot_selected_action.triggered.connect(lambda: self.plot_chart(self.BP12_SSC_diagram_chart, self.selections, False))
        self.BP12_SSC_append_selected_action = self.BP12_SSC_diagram_menu.addAction(self.tr("Append Selected Samples"))
        self.BP12_SSC_append_selected_action.triggered.connect(lambda: self.plot_chart(self.BP12_SSC_diagram_chart, self.selections, True))
        self.BP12_SSC_plot_all_action = self.BP12_SSC_diagram_menu.addAction(self.tr("Plot All Samples"))
        self.BP12_SSC_plot_all_action.triggered.connect(lambda: self.plot_chart(self.BP12_SSC_diagram_chart, self.__dataset.samples, False))


        self.save_action = self.menu.addAction(qta.icon("mdi.microsoft-excel"), self.tr("Save Summary"))
        self.save_action.triggered.connect(self.on_save_clicked)
        self.data_table.customContextMenuRequested.connect(self.show_menu)

    def show_menu(self, pos):
        self.menu.popup(QCursor.pos())

    def show_message(self, title: str, message: str):
        self.msg_box.setWindowTitle(title)
        self.msg_box.setText(message)
        self.msg_box.exec_()

    def show_info(self, message: str):
        self.show_message(self.tr("Info"), message)

    def show_warning(self, message: str):
        self.show_message(self.tr("Warning"), message)

    def show_error(self, message: str):
        self.show_message(self.tr("Error"), message)

    def load_dataset(self):
        self.load_dataset_dialog.show()

    def on_data_loaded(self, dataset: GrainSizeDataset):
        self.__dataset = dataset
        self.current_page_combo_box.clear()
        page_count, left = divmod(self.__dataset.n_samples, self.PAGE_ROWS)
        if left != 0:
            page_count += 1
        self.current_page_combo_box.addItems([f"{self.tr('Page')} {i+1}" for i in range(page_count)])
        self.update_page(0)

    @property
    def is_geometric(self) -> bool:
        return self.geometric_checkbox.isChecked()

    def on_is_geometric_changed(self, state):
        if state == Qt.Checked:
            self.geometric_checkbox.setText(self.tr("Geometric"))
        else:
            self.geometric_checkbox.setText(self.tr("Logarithmic"))
        self.update_page(self.page_index)

    @property
    def use_FW57(self) -> bool:
        return self.FW57_checkbox.isChecked()

    def on_use_FW57_changed(self, state):
        if state == Qt.Checked:
            self.FW57_checkbox.setText(self.tr("Folk and Ward (1957) method"))
        else:
            self.FW57_checkbox.setText(self.tr("Method of moments"))
        self.update_page(self.page_index)

    @property
    def proportion(self) -> str:
        index = self.proportion_combo_box.currentIndex()
        key, description = self.supported_proportions[index]
        return key, description

    @property
    def page_index(self) -> int:
        return self.current_page_combo_box.currentIndex()

    @property
    def n_pages(self) -> int:
        return self.current_page_combo_box.count()

    @property
    def unit(self) -> str:
        return "μm" if self.is_geometric else "φ"

    def update_page(self, page_index: int):
        if self.__dataset is None:
            return

        def write(row: int, col: int, value: str):
            if isinstance(value, str):
                pass
            elif isinstance(value, int):
                value = str(value)
            elif isinstance(value, float):
                value = f"{value: 0.2f}"
            else:
                value = value.__str__()
            item = QTableWidgetItem(value)
            item.setTextAlignment(Qt.AlignCenter)
            self.data_table.setItem(row, col, item)
        # necessary to clear
        self.data_table.clear()
        if page_index == self.n_pages - 1:
            start = page_index * self.PAGE_ROWS
            end = self.__dataset.n_samples
        else:
            start, end = page_index * self.PAGE_ROWS, (page_index+1) * self.PAGE_ROWS
        proportion_key, proportion_desciption = self.proportion
        moment_names = [f"{self.tr('Mean')}[{self.unit}]",
                        self.tr("Mean Desc."),
                        f"{self.tr('Median')} [{self.unit}]",
                        f"{self.tr('Modes')} [{self.unit}]",
                        self.tr("STD (Sorting)"),
                        self.tr("Sorting Desc."),
                        self.tr("Skewness"),
                        self.tr("Skew. Desc."),
                        self.tr("Kurtosis"),
                        self.tr("Kurt. Desc."),
                        f"({proportion_desciption})\n{self.tr('Proportion')} [%]",
                        self.tr("Textural Group\n(Folk, 1954)"),
                        self.tr("Textural Group\nSymbol (Blott & Pye, 2012)"),
                        self.tr("Textural Group\n(Blott & Pye, 2012)")]
        moment_keys = ["mean",
                       "mean_description",
                       "median",
                       "modes",
                       "std",
                       "std_description",
                       "skewness",
                       "skewness_description",
                       "kurtosis",
                       "kurtosis_description",
                        proportion_key,
                       "textural_group_Folk54",
                       "textural_group_BP12_symbol",
                       "textural_group_BP12"]
        self.data_table.setRowCount(end-start)
        self.data_table.setColumnCount(len(moment_names))
        self.data_table.setHorizontalHeaderLabels(moment_names)
        self.data_table.setVerticalHeaderLabels([sample.name for sample in self.__dataset.samples[start: end]])
        for row, sample in enumerate(self.__dataset.samples[start: end]):
            geometric_moments, logarithmic_moments = get_moments(sample.classes_μm, sample.classes_φ, sample.distribution, FW57=self.use_FW57)
            if self.is_geometric:
                moments = geometric_moments
            else:
                moments = logarithmic_moments
            for col, key in enumerate(moment_keys):
                if key == "modes":
                    write(row, col, ", ".join([f"{m:0.2f}" for m in moments[key]]))
                elif key[-11:] == "_proportion":
                    write(row, col, ", ".join([f"{p*100:0.2f}" for p in moments[key]]))
                else:
                    write(row, col, moments[key])

        self.data_table.resizeColumnsToContents()

    @property
    def selections(self):
        if self.__dataset.n_samples == 0:
            self.show_warning(self.tr("Dataset has not been loaded."))
            return []

        start = self.page_index*self.PAGE_ROWS
        temp = set()
        for item in self.data_table.selectedRanges():
            for i in range(item.topRow(), min(self.PAGE_ROWS+1, item.bottomRow()+1)):
                temp.add(i+start)
        indexes = list(temp)
        indexes.sort()
        samples = [self.__dataset.samples[i] for i in indexes]
        return samples

    def on_previous_button_clicked(self):
        if self.page_index > 0:
            self.current_page_combo_box.setCurrentIndex(self.page_index-1)

    def on_next_button_clicked(self):
        if self.page_index < self.n_pages - 1:
            self.current_page_combo_box.setCurrentIndex(self.page_index+1)

    def plot_chart(self, chart, samples, append):
        if len(samples) == 0:
            return
        chart.show_samples(samples, append=append)
        chart.show()

    def save_file(self, filename: str):
        wb = openpyxl.Workbook()
        prepare_styles(wb)

        ws = wb.active
        ws.title = self.tr("README")
        description = \
            """
            This Excel file was generated by QGrain ({0}).

            It contanins one sheet:
            1. The sheet puts the statistic parameters and the textural group of the samples.

            The statistic formulas are referred to Blott & Pye (2001)'s work.
            The classification of textural groups is referred to Folk (1957) and Blott & Pye (2012)'s scheme.

            References:
                1.Blott, S. J. & Pye, K. Particle size scales and classification of sediment types based on particle size distributions: Review and recommended procedures. Sedimentology 59, 2071–2096 (2012).
                2.Blott, S. J. & Pye, K. GRADISTAT: a grain size distribution and statistics package for the analysis of unconsolidated sediments. Earth Surf. Process. Landforms 26, 1237–1248 (2001).
                3.Folk, R. L. The Distinction between Grain Size and Mineral Composition in Sedimentary-Rock Nomenclature. The Journal of Geology 62, 344–359 (1954).

            """.format(QGRAIN_VERSION)

        def write(row, col, value, style="normal_light"):
            cell = ws.cell(row+1, col+1, value=value)
            cell.style = style

        lines_of_desc = description.split("\n")
        for row, line in enumerate(lines_of_desc):
            write(row, 0, line, style="description")
        ws.column_dimensions[column_to_char(0)].width = 200

        ws = wb.create_sheet(self.tr("Moments"))
        proportion_key, proportion_desciption = self.proportion
        moment_names = [f"{self.tr('Mean')}[{self.unit}]",
                        self.tr("Mean Desc."),
                        f"{self.tr('Median')} [{self.unit}]",
                        f"{self.tr('Modes')} [{self.unit}]",
                        self.tr("STD (Sorting)"),
                        self.tr("Sorting Desc."),
                        self.tr("Skewness"),
                        self.tr("Skew. Desc."),
                        self.tr("Kurtosis"),
                        self.tr("Kurt. Desc."),
                        f"({proportion_desciption})\n{self.tr('Proportion')} [%]",
                        self.tr("Textural Group\n(Folk, 1954)"),
                        self.tr("Textural Group\nSymbol (Blott & Pye, 2012)"),
                        self.tr("Textural Group\n(Blott & Pye, 2012)")]
        moment_keys = ["mean",
                       "mean_description",
                       "median",
                       "modes",
                       "std",
                       "std_description",
                       "skewness",
                       "skewness_description",
                       "kurtosis",
                       "kurtosis_description",
                        proportion_key,
                       "textural_group_Folk54",
                       "textural_group_BP12_symbol",
                       "textural_group_BP12"]
        write(0, 0, self.tr("Sample Name"), style="header")
        ws.column_dimensions[column_to_char(0)].width = 16
        for col, moment_name in enumerate(moment_names, 1):
            write(0, col, moment_name, style="header")
            if col in (2, 4, 6, 8, 10, 11, 12, 14):
                ws.column_dimensions[column_to_char(col)].width = 30
            else:
                ws.column_dimensions[column_to_char(col)].width = 16
        ws.column_dimensions[column_to_char(len(moment_names))].width = 40
        for row, sample in enumerate(self.__dataset.samples, 1):
            if row % 2 == 0:
                style = "normal_dark"
            else:
                style = "normal_light"
            write(row, 0, sample.name, style=style)
            geometric_moments, logarithmic_moments = get_moments(sample.classes_μm, sample.classes_φ, sample.distribution, FW57=self.use_FW57)
            if self.is_geometric:
                moments = geometric_moments
            else:
                moments = logarithmic_moments
            for col, key in enumerate(moment_keys, 1):
                if key == "modes":
                    write(row, col, ", ".join([f"{m:0.4f}" for m in moments[key]]), style=style)
                elif key[-11:] == "_proportion":
                    write(row, col, ", ".join([f"{p*100:0.4f}" for p in moments[key]]), style=style)
                else:
                    write(row, col, moments[key], style=style)

        wb.save(filename)
        wb.close()

    def on_save_clicked(self):
        if self.__dataset is None or self.__dataset.n_samples == 0:
            self.show_warning(self.tr("Dataset has not been loaded."))
            return

        filename, _  = self.file_dialog.getSaveFileName(self, self.tr("Select Filename"),
                                         None, "Excel (*.xlsx)")
        if filename is None or filename == "":
            return

        try:
            self.save_file(filename)
            self.show_info(self.tr("The summary of this dataset has been saved to:\n    {0}").format(filename))
        except Exception as e:
            self.show_error(self.tr("Error raised while save summary to Excel file.\n    {0}").format(e.__str__()))

if __name__ == "__main__":
    import sys
    from QGrain.entry import setup_app
    app = setup_app()
    main = GrainSizeDatasetViewer()
    main.show()
    sys.exit(app.exec_())
