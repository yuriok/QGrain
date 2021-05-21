__all__ = ["HCResolverPanel"]

import typing

import numpy as np
import openpyxl
import qtawesome as qta
from PySide2.QtCore import Qt, QCoreApplication
from PySide2.QtWidgets import (QDialog, QFileDialog, QGridLayout, QLabel, QComboBox,QApplication,
                               QMessageBox, QPushButton, QSpinBox)
from QGrain.charts.HierarchicalChart import HierarchicalChart
from QGrain.models.GrainSizeDataset import GrainSizeDataset
from QGrain.ui.LoadDatasetDialog import LoadDatasetDialog
from QGrain.use_excel import column_to_char, prepare_styles

from scipy.cluster.hierarchy import fcluster, fclusterdata, linkage
from QGrain.charts.FrequencyCurveChart import FrequencyCurveChart
from QGrain import QGRAIN_VERSION

class HCResolverPanel(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent=parent, f=Qt.Window)
        self.setWindowTitle(self.tr("HC Resolver"))
        self.init_ui()
        self.frequency_chart = FrequencyCurveChart(parent=self, toolbar=True)
        self.load_dataset_dialog = LoadDatasetDialog(parent=self)
        self.load_dataset_dialog.dataset_loaded.connect(self.on_dataset_loaded)
        self.file_dialog = QFileDialog(self)
        self.msg_box = QMessageBox(self)
        self.msg_box.setWindowFlags(Qt.Drawer)
        self.__dataset = None
        self.__X = None
        self.__last_result = None

    def init_ui(self):
        self.main_layout = QGridLayout(self)
        self.chart = HierarchicalChart(toolbar=True)
        self.main_layout.addWidget(self.chart, 0, 0, 1, 4)

        self.load_dataset_button = QPushButton(qta.icon("fa5s.database"), self.tr("Load Dataset"))
        self.load_dataset_button.clicked.connect(lambda: self.load_dataset_dialog.show())
        self.main_layout.addWidget(self.load_dataset_button, 1, 0, 1, 2)

        self.n_samples_label = QLabel(self.tr("N<sub>components</sub>"))
        self.n_samples_label.setToolTip(self.tr("The Number of samples in the loaded dataset."))
        self.n_samples_display = QLabel(self.tr("Unknown"))
        self.main_layout.addWidget(self.n_samples_label, 1, 2)
        self.main_layout.addWidget(self.n_samples_display, 1, 3)

        self.supported_methods = [
            "single", "complete",
            "average", "weighted",
            "centroid", "median",
            "ward"]
        self.supported_distances = [
            "braycurtis", "canberra",
            "chebyshev", "cityblock",
            "correlation", "cosine",
            "dice", "euclidean",
            "hamming", "jaccard",
            "jensenshannon", "kulsinski",
            "mahalanobis", "matching",
            "minkowski", "rogerstanimoto",
            "russellrao", "seuclidean",
            "sokalmichener", "sokalsneath",
            "sqeuclidean", "yule"]

        self.linkage_method_label = QLabel(self.tr("Linkage Method"))
        self.linkage_method_label.setToolTip(self.tr("The linkage method for calculating the distance between the newly formed cluster and each observation vector."))
        self.linkage_method_combo_box = QComboBox()
        self.linkage_method_combo_box.addItems(self.supported_methods)
        self.linkage_method_combo_box.setCurrentText("ward")
        self.distance_label = QLabel(self.tr("Distance"))
        self.distance_label.setToolTip(self.tr("The distance metric to use in the case that y is a collection of observation vectors."))
        self.distance_combo_box = QComboBox()
        self.distance_combo_box.addItems(self.supported_distances)
        self.distance_combo_box.setCurrentText("euclidean")
        self.main_layout.addWidget(self.linkage_method_label, 2, 0)
        self.main_layout.addWidget(self.linkage_method_combo_box, 2, 1)
        self.main_layout.addWidget(self.distance_label, 2, 2)
        self.main_layout.addWidget(self.distance_combo_box, 2, 3)

        self.p_label = QLabel(self.tr("p"))
        self.p_label.setToolTip(self.tr("Controls the number of leaves at the bottom level of the figure."))
        self.p_input = QSpinBox()
        self.p_input.setMinimum(1)
        self.main_layout.addWidget(self.p_label, 3, 0)
        self.main_layout.addWidget(self.p_input, 3, 1)
        self.n_clusers_label = QLabel(self.tr("N<sub>clusters</sub>"))
        self.n_clusers_label.setToolTip(self.tr("Controls the number of clusters of this clustering algorithm."))
        self.n_clusers_input = QSpinBox()
        self.n_clusers_input.setMinimum(2)
        self.main_layout.addWidget(self.n_clusers_label, 3, 2)
        self.main_layout.addWidget(self.n_clusers_input, 3, 3)

        self.perform_button = QPushButton(qta.icon("ei.ok-sign"), self.tr("Perform"))
        self.perform_button.setEnabled(False)
        self.perform_button.clicked.connect(self.on_perform_clicked)
        self.show_typical_button = QPushButton(qta.icon("fa.area-chart"), self.tr("Show Typical"))
        self.show_typical_button.clicked.connect(self.on_show_typical_clicked)
        self.show_typical_button.setEnabled(False)
        self.save_button = QPushButton(qta.icon("fa5s.save"), self.tr("Save"))
        self.save_button.setEnabled(False)
        self.show_all_button = QPushButton(qta.icon("mdi.view-grid-outline"), self.tr("Show All"))
        self.show_all_button.setEnabled(False)
        self.show_all_button.clicked.connect(self.on_show_all_clicked)
        self.save_button.clicked.connect(self.on_save_clicked)
        self.main_layout.addWidget(self.perform_button, 4, 0, 1, 2)
        self.main_layout.addWidget(self.show_typical_button, 4, 2, 1, 2)
        self.main_layout.addWidget(self.save_button, 5, 0, 1, 2)
        self.main_layout.addWidget(self.show_all_button, 5, 2, 1, 2)

    def show_message(self, title: str, message: str):
        self.msg_box.setWindowTitle(title)
        self.msg_box.setText(message)
        self.msg_box.exec_()

    def show_info(self, message: str):
        self.show_message(self.tr("Info"), message)

    def show_warning(self, message: str):
        self.show_message(self.tr("Warning"), message)

    def show_error(self, message: str):
        self.show_message(self.tr("Error"), message)

    @property
    def linkage_method(self) -> str:
        return self.linkage_method_combo_box.currentText()

    @property
    def distance(self) -> str:
        return self.distance_combo_box.currentText()

    @property
    def p(self) -> int:
        return self.p_input.value()

    @property
    def n_clusters(self) -> int:
        return self.n_clusers_input.value()

    def on_dataset_loaded(self, dataset: GrainSizeDataset):
        self.__dataset = dataset
        self.__X = dataset.X
        if dataset.n_samples == 0:
            self.show_error("Dataset is empty.")
            return

        self.perform_button.setEnabled(True)
        self.n_samples_display.setText(str(dataset.n_samples))
        self.p_input.setMaximum(dataset.n_samples)
        self.n_clusers_input.setMaximum(dataset.n_samples-1)
        if dataset.n_samples > 20:
            self.p_input.setValue(10)

    @property
    def n_components(self):
        return self.n_samples_display.value()

    def on_perform_clicked(self):
        try:
            linkage_matrix = linkage(self.__X, method=self.linkage_method, metric=self.distance)
        except ValueError as e:
            self.show_error(self.tr("The linkage method is not compatible with the distance.\n    {0}").format(e.__str__()))
            return
        dendrogram_res = self.chart.show_result(linkage_matrix, p=self.p)
        self.__last_result = (self.__dataset, self.__X, linkage_matrix, dendrogram_res)
        self.show_typical_button.setEnabled(True)
        self.show_all_button.setEnabled(True)
        self.save_button.setEnabled(True)

    def on_show_typical_clicked(self):
        dataset, X, linkage_matrix, dendrogram_res = self.__last_result
        flags = fcluster(linkage_matrix, self.n_clusters, criterion="maxclust")
        samples = []
        flag_set = set()
        for i, flag in enumerate(flags):
            if len(flag_set) == self.n_clusters:
                break

            if flag not in flag_set:
                samples.append(dataset.samples[i])
                flag_set.add(flag)

        self.frequency_chart.show_samples(samples, append=False, title=self.tr("Typical Samples of All Clusters"))
        self.frequency_chart.show()

    def on_show_all_clicked(self):
        if self.n_clusters > 20:
            self.show_warning(self.tr("N_clusters is greater than 20, it will generate too much windows on your screen."))

        dataset, X, linkage_matrix, dendrogram_res = self.__last_result
        flags = fcluster(linkage_matrix, self.n_clusters, criterion="maxclust")
        flag_set = set(flags)
        rect = QApplication.primaryScreen().availableGeometry()
        x0 = rect.x()
        y0 = rect.y()
        max_w = rect.width()
        max_h = rect.height()
        span = 20
        epoch = 0
        w = 400
        h = 300
        x = x0 + epoch % 10 * 20 + span
        y = y0 + epoch % 10 * 20 + span
        span = 20
        for flag in flag_set:
            samples = []
            for sample, in_this_cluster in zip(dataset.samples, np.equal(flags, flag)):
                if in_this_cluster:
                    samples.append(sample)
            chart = FrequencyCurveChart(parent=self, toolbar=True)
            chart.show_samples(samples, append=False, title=self.tr("Cluster{0}").format(flag))
            chart.setGeometry(x, y, w, h)
            x += w + span
            if x > max_w - w:
                x = x0 + epoch % 10 * 40 + span
                y+= h + span
            if y > max_h - h:
                epoch += 1
                x = x0 + epoch % 10 * 40 + span
                y = x0 + epoch % 10 * 40 + span
            chart.show()
            QCoreApplication.processEvents()

    def on_save_clicked(self):
        if self.__last_result is None:
            self.show_warning(self.tr("The hierarchy clustering algorithm has not been performed."))
            return
        filename, _ = self.file_dialog.getSaveFileName(
            None, self.tr("Choose a filename to save the hierarchy clustering result"),
            None, "Microsoft Excel (*.xlsx)")
        if filename is None or filename == "":
            return
        try:
            self.save_as_xlsx(filename)
            self.show_info(self.tr("The hierarchy clustering result has been saved to:\n    {0}").format(filename))
        except Exception as e:
            self.show_error(self.tr("Error raised while save hierarchy clustering result to Excel file.\n    {0}").format(e.__str__()))

    def save_as_xlsx(self, filename: str):
        assert self.__last_result is not None
        dataset, X, linkage_matrix, dendrogram_res = self.__last_result
        n_clusters = self.n_clusters
        flags = fcluster(linkage_matrix, n_clusters, criterion="maxclust")
        flag_set = set(flags)
        n_clusters = len(flag_set)

        wb = openpyxl.Workbook()
        prepare_styles(wb)
        ws = wb.active
        ws.title = self.tr("README")
        description = \
            """
            This Excel file was generated by QGrain ({0}).

            It contanins three (or n_clusters + 3) sheets:
            1. The first sheet is the dataset which was used to perform the hierarchy clustering algorithm.
            2. The second sheet is used to put the clustering flags of all samples.
            3. The third sheet is the typical sampels (i.e, the first sample of each cluster was selected).
            4. If the number of clusters less equal to 100, the samples of each cluster will be save to individual sheets.

            The base hierarchy clusrting algorithm is implemented by Scipy. You can get the details of algorithm from the following website.
            https://docs.scipy.org/doc/scipy/reference/cluster.hierarchy.html

            """.format(QGRAIN_VERSION)

        def write(row, col, value, style="normal_light"):
            cell = ws.cell(row+1, col+1, value=value)
            cell.style = style

        lines_of_desc = description.split("\n")
        for row, line in enumerate(lines_of_desc):
            write(row, 0, line, style="description")
        ws.column_dimensions[column_to_char(0)].width = 200

        ws = wb.create_sheet(self.tr("Dataset"))
        write(0, 0, self.tr("Sample Name"), style="header")
        ws.column_dimensions[column_to_char(0)].width = 16
        for col, value in enumerate(dataset.classes_μm, 1):
            write(0, col, value, style="header")
            ws.column_dimensions[column_to_char(col)].width = 10
        for row, sample in enumerate(dataset.samples, 1):
            if row % 2 == 0:
                style = "normal_dark"
            else:
                style = "normal_light"
            write(row, 0, sample.name, style=style)
            for col, value in enumerate(sample.distribution, 1):
                write(row, col, value, style=style)
            QCoreApplication.processEvents()

        ws = wb.create_sheet(self.tr("Flags"))
        write(0, 0, self.tr("Sample Name"), style="header")
        write(0, 1, self.tr("Flag of the Cluster"), style="header")
        ws.column_dimensions[column_to_char(0)].width = 16
        ws.column_dimensions[column_to_char(1)].width = 16
        for row, (sample, flag) in enumerate(zip(dataset.samples, flags), 1):
            if row % 2 == 0:
                style = "normal_dark"
            else:
                style = "normal_light"
            write(row, 0, sample.name, style=style)
            write(row, 1, flag, style=style)
            QCoreApplication.processEvents()

        typical_samples = []
        temp_flag_set = set()
        for i, flag in enumerate(flags):
            if len(temp_flag_set) == n_clusters:
                break
            if flag not in temp_flag_set:
                typical_samples.append(dataset.samples[i])
                temp_flag_set.add(flag)

        ws = wb.create_sheet(self.tr("Typical Samples"))
        write(0, 0, self.tr("Sample Name"), style="header")
        ws.column_dimensions[column_to_char(0)].width = 16
        for col, value in enumerate(dataset.classes_μm, 1):
            write(0, col, value, style="header")
            ws.column_dimensions[column_to_char(col)].width = 10
        for row, sample in enumerate(typical_samples, 1):
            if row % 2 == 0:
                style = "normal_dark"
            else:
                style = "normal_light"
            write(row, 0, sample.name, style=style)
            for col, value in enumerate(sample.distribution, 1):
                write(row, col, value, style=style)
            QCoreApplication.processEvents()

        if n_clusters <= 100:
            for flag in flag_set:
                samples = []
                for sample, in_this_cluster in zip(dataset.samples, np.equal(flags, flag)):
                    if in_this_cluster:
                        samples.append(sample)

                ws = wb.create_sheet(self.tr("Cluster{0}").format(flag))
                write(0, 0, self.tr("Sample Name"), style="header")
                ws.column_dimensions[column_to_char(0)].width = 16
                for col, value in enumerate(dataset.classes_μm, 1):
                    write(0, col, value, style="header")
                    ws.column_dimensions[column_to_char(col)].width = 10
                for row, sample in enumerate(samples, 1):
                    if row % 2 == 0:
                        style = "normal_dark"
                    else:
                        style = "normal_light"
                    write(row, 0, sample.name, style=style)
                    for col, value in enumerate(sample.distribution, 1):
                        write(row, col, value, style=style)
                    QCoreApplication.processEvents()

        wb.save(filename)
        wb.close()


if __name__ == "__main__":
    import sys

    from QGrain.entry import setup_app
    app, splash = setup_app()
    main = HCResolverPanel()
    main.show()
    splash.finish(main)
    sys.exit(app.exec_())
