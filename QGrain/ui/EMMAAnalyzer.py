__all__ = ["EMMAAnalyzer"]

import os
import pickle
import typing

import numpy as np
import openpyxl
from PySide6.QtCore import QCoreApplication, Qt
from PySide6.QtWidgets import (QCheckBox, QComboBox, QDialog, QFileDialog,
                               QGridLayout, QGroupBox, QLabel, QListWidget,
                               QMessageBox, QProgressBar, QPushButton,
                               QSpinBox)

from .. import QGRAIN_VERSION
from ..chart.EMMAResultChart import EMMAResultChart
from ..chart.EMMASummaryChart import EMMASummaryChart
from ..emma import EMMAResolver, EMMAResolverSetting, EMMAResult, KernelType
from ..io import column_to_char, prepare_styles
from ..model import GrainSizeDataset
from ..statistic import convert_μm_to_φ
from .EMMASettingDialog import EMMASettingDialog
from ..ui.LoadDatasetDialog import LoadDatasetDialog


class EMMAAnalyzer(QDialog):
    SUPPORT_KERNELS = (
        KernelType.Nonparametric,
        KernelType.Normal,
        KernelType.SkewNormal,
        KernelType.Weibull,
        KernelType.GeneralWeibull)
    def __init__(self, parent=None):
        super().__init__(parent=parent, f=Qt.Window)
        self.setWindowTitle(self.tr("EMMA Resolver"))
        self.init_ui()
        self.normal_msg = QMessageBox(self)
        self.__dataset = None # type: GrainSizeDataset
        self.__initial_params = None # type: np.ndarray
        self.__result_list = [] # type: list[EMMAResult]
        self.resolver_setting = EMMASettingDialog(parent=self)
        self.resolver_setting.setting = EMMAResolverSetting(min_niter=800, max_niter=1200, tol=-6, ftol=1e-8, lr=5e-2)
        self.file_dialog = QFileDialog(parent=self)
        self.emma_result_chart = EMMAResultChart(toolbar=True)
        self.emma_summary_chart = EMMASummaryChart(toolbar=True)

    def init_ui(self):
        self.setAttribute(Qt.WA_StyledBackground, True)
        self.main_layout = QGridLayout(self)
        # Control group
        self.control_group = QGroupBox(self.tr("Control"))
        self.control_layout = QGridLayout(self.control_group)
        self.main_layout.addWidget(self.control_group, 0, 0)
        self.load_dataset_button = QPushButton(self.tr("Load Dataset"))
        self.control_layout.addWidget(self.load_dataset_button, 0, 0, 1, 2)
        self.configure_button = QPushButton(self.tr("Configure Algorithm"))
        self.configure_button.clicked.connect(self.on_configure_clicked)
        self.control_layout.addWidget(self.configure_button, 1, 0, 1, 2)
        self.n_samples_label = QLabel(self.tr("Number of Samples"))
        self.n_samples_display = QLabel(self.tr("Unknown"))
        self.control_layout.addWidget(self.n_samples_label, 2, 0)
        self.control_layout.addWidget(self.n_samples_display, 2, 1)
        self.kernel_label = QLabel(self.tr("Kernel Type"))
        self.kernel_combo_box = QComboBox()
        self.kernel_combo_box.addItems([kernel_type.value for kernel_type in self.SUPPORT_KERNELS])
        self.kernel_combo_box.currentIndexChanged.connect(self.on_kernel_type_changed)
        self.control_layout.addWidget(self.kernel_label, 3, 0)
        self.control_layout.addWidget(self.kernel_combo_box, 3, 1)
        self.n_members_label = QLabel("Number of Members")
        self.n_members_input = QSpinBox()
        self.n_members_input.setRange(1, 10)
        self.n_members_input.valueChanged.connect(self.on_n_members_changed)
        self.control_layout.addWidget(self.n_members_label, 4, 0)
        self.control_layout.addWidget(self.n_members_input, 4, 1)
        self.configure_initial_button = QPushButton(self.tr("Configure Initial EMs"))
        self.configure_initial_button.clicked.connect(self.on_configure_initial_clicked)
        self.configure_initial_button.setEnabled(False)
        self.update_EMs_checkbox = QCheckBox(self.tr("Update EM Distributions"))
        self.update_EMs_checkbox.setChecked(True)
        self.perform_button = QPushButton(self.tr("Perform"))
        self.perform_button.clicked.connect(self.on_perform_clicked)
        self.perform_button.setEnabled(False)
        self.control_layout.addWidget(self.configure_initial_button, 5, 0, 1, 2)
        self.control_layout.addWidget(self.update_EMs_checkbox, 6, 0, 1, 2)
        self.control_layout.addWidget(self.perform_button, 7, 0, 1, 2)
        # Result group
        self.result_group = QGroupBox(self.tr("Result"))
        self.result_layout = QGridLayout(self.result_group)
        self.main_layout.addWidget(self.result_group, 0, 1)
        self.result_list_widget = QListWidget()
        self.result_layout.addWidget(self.result_list_widget, 0, 0, 1, 2)
        self.remove_result_button = QPushButton(self.tr("Remove"))
        self.remove_result_button.clicked.connect(self.on_remove_clicked)
        self.show_result_button = QPushButton(self.tr("Show"))
        self.show_result_button.clicked.connect(self.on_show_clicked)
        self.load_dump_button = QPushButton(self.tr("Load Dump"))
        self.load_dump_button.clicked.connect(self.on_load_dump_clicked)
        self.save_button = QPushButton(self.tr("Save"))
        self.save_button.clicked.connect(self.on_save_clicked)
        self.remove_result_button.setEnabled(False)
        self.show_result_button.setEnabled(False)
        self.save_button.setEnabled(False)
        self.result_layout.addWidget(self.remove_result_button, 1, 0)
        self.result_layout.addWidget(self.show_result_button, 1, 1)
        self.result_layout.addWidget(self.load_dump_button, 2, 0)
        self.result_layout.addWidget(self.save_button, 2, 1)

    def show_message(self, title: str, message: str):
        self.normal_msg.setWindowTitle(title)
        self.normal_msg.setText(message)
        self.normal_msg.exec_()

    def show_info(self, message: str):
        self.show_message(self.tr("Info"), message)

    def show_warning(self, message: str):
        self.show_message(self.tr("Warning"), message)

    def show_error(self, message: str):
        self.show_message(self.tr("Error"), message)

    @property
    def kernel_type(self) -> KernelType:
        kernel_type = self.SUPPORT_KERNELS[self.kernel_combo_box.currentIndex()]
        return kernel_type

    @property
    def n_members(self):
        return self.n_members_input.value()

    @property
    def n_results(self) -> int:
        return len(self.__result_list)

    @property
    def selected_index(self):
        indexes = self.result_list_widget.selectedIndexes()
        if len(indexes) == 0:
            return 0
        else:
            return indexes[0].row()

    @property
    def selected_result(self):
        if self.n_results > 0:
            return self.__result_list[self.selected_index]

    def on_dataset_loaded(self, dataset: GrainSizeDataset):
        self.__dataset = dataset
        self.n_samples_display.setText(str(self.__dataset.n_samples))
        self.perform_button.setEnabled(True)
        self.configure_initial_button.setEnabled(True)

    def on_configure_clicked(self):
        self.resolver_setting.show()

    def on_configure_initial_clicked(self):
        if self.__dataset is None:
            self.show_error(self.tr("Dataset has not been loaded."))
            return

        filename, _ = self.file_dialog.getOpenFileName(
            self, self.tr("Choose a excel file which contains the end member distributions at the first sheet"),
            None, f"{self.tr('Microsoft Excel')} (*.xlsx)")
        if filename is None or filename == "":
            return
        try:
            wb = openpyxl.load_workbook(filename, read_only=True, data_only=True)
            ws = wb[wb.sheetnames[0]]
            raw_data = [[value for value in row] for row in ws.values]
            classes_μm = np.array(raw_data[0][1:], dtype=np.float64)
            classes_φ = convert_μm_to_φ(classes_μm)
            distributions = [np.array(row[1:], dtype=np.float64) for row in raw_data[1:]]
        except Exception as e:
            self.show_error(self.tr("Error raised while loading end member distributions.\n    {0}").format(e.__str__()))
            return
        if len(classes_μm) < 10:
            self.show_error(self.tr("The number of grain size classes is too less."))
            return
        for i in range(len(classes_μm)-1):
            if classes_μm[i+1] <= classes_μm[i]:
                self.show_error(self.tr("The grain size classes is not incremental, please check the first row."))
                return
        if np.any(np.isnan(classes_μm)):
            self.show_error(self.tr("There is at least one NaN value in grain size classes."))
            return
        if len(distributions) > 10:
            self.show_error(self.tr("There are more than 10 rows representing the end member distributions, please check."))
            return
        for distribution in distributions:
            if len(classes_μm) != len(distribution):
                self.show_error(self.tr("Some distributions have different numbers of grain size classes."))
                return
            if np.any(np.isnan(distribution)):
                self.show_error(self.tr("There is at least one NaN value in the end member distributions."))
                return
            if abs(np.sum(distribution) - 1.0) > 0.05:
                self.show_error(self.tr("The sum of one distribution is not equal to 1."))
                return

        self.n_members_input.setValue(len(distributions))
        resolver = EMMAResolver()
        params = resolver.get_params(self.__dataset, classes_φ, distributions, self.kernel_type)
        params = params.astype(np.float32)
        self.__initial_params = params
        self.show_info(self.tr("The initial end member distributions have been appointed. The number of end members has been set to [{0}].").format(len(distributions)))

    def on_kernel_type_changed(self, index: int):
        if self.__initial_params is not None:
            self.__initial_params = None
            self.show_warning(self.tr("The appointed initial end members will be discarded."))

    def on_n_members_changed(self, n_member: int):
        if self.__initial_params is not None:
            self.__initial_params = None
            self.show_warning(self.tr("The appointed initial end members will be discarded."))

    def on_perform_clicked(self):
        if self.__dataset is None:
            self.show_error(self.tr("Dataset has not been loaded."))
            return

        self.perform_button.setEnabled(False)
        self.configure_initial_button.setEnabled(False)
        resolver = EMMAResolver()
        resolver_setting = self.resolver_setting.setting
        update_end_members = self.update_EMs_checkbox.isChecked()
        result = resolver.try_fit(self.__dataset, self.kernel_type, self.n_members, resolver_setting, self.__initial_params, update_end_members)
        self.add_results([result])
        self.perform_button.setEnabled(True)
        self.configure_initial_button.setEnabled(True)

    def get_result_name(self, result: EMMAResult):
        if self.update_EMs_checkbox.isChecked():
            fixed_str = " "
        else:
            fixed_str = " Fixed "
        return f"{result.n_members}{fixed_str}{result.kernel_type.value}"

    def add_results(self, results: typing.List[EMMAResult]):
        if self.n_results == 0:
            self.remove_result_button.setEnabled(True)
            self.show_result_button.setEnabled(True)
            self.save_button.setEnabled(True)

        self.__result_list.extend(results)
        self.result_list_widget.addItems([self.get_result_name(result) for result in results])

    def on_remove_clicked(self):
        if self.n_results == 0:
            return
        else:
            index = self.selected_index
            self.__result_list.pop(index)
            self.result_list_widget.takeItem(index)

        if self.n_results == 0:
            self.remove_result_button.setEnabled(False)
            self.show_result_button.setEnabled(False)
            self.save_button.setEnabled(False)

    def on_show_clicked(self):
        result = self.selected_result
        if result is None:
            return
        else:
            self.emma_result_chart.show_result(result)
            self.emma_result_chart.show()

    def on_load_dump_clicked(self):
        filename, _  = self.file_dialog.getOpenFileName(self, self.tr("Select the dump file of the EMMA result(s)"),
                                         None, f"{self.tr('Binary Dump')} (*.dump)")
        if filename is None or filename == "":
            return
        with open(filename, "rb") as f:
            results = pickle.load(f)
            invalid = False
            if isinstance(results, list):
                for result in results:
                    if not isinstance(result, EMMAResult):
                        invalid = True
                        break
            else:
                invalid = True
            if invalid:
                self.show_error(self.tr("The dump file does not contain any EMMA result."))
                return
            else:
                self.add_results(results)

    def save_result_excel(self, filename: str, result: EMMAResult):
        # get the mode size of each end-members
        modes = [(i, result.dataset.classes_μm[np.unravel_index(np.argmax(result.end_members[i]), result.end_members[i].shape)]) for i in range(result.n_members)]
        # sort them by mode size
        modes.sort(key=lambda x: x[1])
        wb = openpyxl.Workbook()
        prepare_styles(wb)

        ws = wb.active
        ws.title = self.tr("README")
        description = \
            """
            This Excel file was generated by QGrain ({0}).

            Please cite:
            Liu, Y., Liu, X., Sun, Y., 2021. QGrain: An open-source and easy-to-use software for the comprehensive analysis of grain size distributions. Sedimentary Geology 423, 105980. https://doi.org/10.1016/j.sedgeo.2021.105980

            It contanins three sheets:
            1. The first sheet is the dataset which was used to perform the EMMA algorithm.
            2. The second sheet is used to put the distributions of all end-members.
            3. The third sheet is the end-member fractions of all samples.

            This EMMA algorithm was implemented by QGrian, using the famous machine learning framework, PyTorch.

            EMMA algorithm details
                N_samples: {1},
                Distribution Type: {2},
                N_members: {3},
                N_iterations: {4},
                Spent Time: {5} s,

                Computing Device: {6},
                Distance: {7},
                Minimum N_iterations: {8},
                Maximum N_iterations: {9},
                Learning Rate: {10},
                eps: {11},
                tol: {12},
                ftol: {13}

            """.format(QGRAIN_VERSION,
                    result.dataset.n_samples,
                    result.kernel_type.name,
                    result.n_members,
                    result.n_iterations,
                    result.time_spent,
                    result.resolver_setting.device,
                    result.resolver_setting.distance,
                    result.resolver_setting.min_niter,
                    result.resolver_setting.max_niter,
                    result.resolver_setting.lr,
                    result.resolver_setting.eps,
                    result.resolver_setting.tol,
                    result.resolver_setting.ftol)

        def write(row, col, value, style="normal_light"):
            cell = ws.cell(row+1, col+1, value=value)
            cell.style = style

        lines_of_desc = description.split("\n")
        for row, line in enumerate(lines_of_desc):
            write(row, 0, line, style="description")
        ws.column_dimensions[column_to_char(0)].width = 200

        ws = wb.create_sheet(self.tr("Dataset"))
        write(0, 0, self.tr("Sample Name"), style="header")
        ws.column_dimensions[column_to_char(0)].width = 16
        for col, value in enumerate(result.dataset.classes_μm, 1):
            write(0, col, value, style="header")
            ws.column_dimensions[column_to_char(col)].width = 10
        for row, sample in enumerate(result.dataset.samples, 1):
            if row % 2 == 0:
                style = "normal_dark"
            else:
                style = "normal_light"
            write(row, 0, sample.name, style=style)
            for col, value in enumerate(sample.distribution, 1):
                write(row, col, value, style=style)
            QCoreApplication.processEvents()

        ws = wb.create_sheet(self.tr("End-members"))
        write(0, 0, self.tr("End-member"), style="header")
        ws.column_dimensions[column_to_char(0)].width = 16
        for col, value in enumerate(result.dataset.classes_μm, 1):
            write(0, col, value, style="header")
            ws.column_dimensions[column_to_char(col)].width = 10
        for row, (index, _) in enumerate(modes, 1):
            if row % 2 == 0:
                style = "normal_dark"
            else:
                style = "normal_light"
            write(row, 0, f"EM{row}", style=style)
            for col, value in enumerate(result.end_members[index], 1):
                write(row, col, value, style=style)
            QCoreApplication.processEvents()

        ws = wb.create_sheet(self.tr("Fractions"))
        write(0, 0, self.tr("Sample Name"), style="header")
        ws.column_dimensions[column_to_char(0)].width = 16
        for i in range(result.n_members):
            write(0, i+1, f"EM{i+1}", style="header")
            ws.column_dimensions[column_to_char(i+1)].width = 10
        for row, fractions in enumerate(result.proportions, 1):
            if row % 2 == 0:
                style = "normal_dark"
            else:
                style = "normal_light"
            write(row, 0, result.dataset.samples[row-1].name, style=style)
            for col, (index, _) in enumerate(modes, 1):
                write(row, col, fractions[index], style=style)
            QCoreApplication.processEvents()

        wb.save(filename)
        wb.close()

    def on_save_clicked(self):
        if self.n_results == 0:
            self.show_warning(self.tr("There is not an EMMA result in the list."))
            return

        filename, _ = self.file_dialog.getSaveFileName(
            self, self.tr("Choose a filename to save the EMMA result(s) in list"),
            None, f"{self.tr('Binary Dump')} (*.dump);;{self.tr('Microsoft Excel')} (*.xlsx)")
        if filename is None or filename == "":
            return
        _, ext = os.path.splitext(filename)

        if ext == ".dump":
            with open(filename, "wb") as f:
                pickle.dump(self.__result_list, f)
                self.show_info(self.tr("All results in list has been saved."))
        elif ext == ".xlsx":
            try:
                result = self.selected_result
                self.save_result_excel(filename, result)
                self.show_info(self.tr("The selected result has been saved."))
            except Exception as e:
                self.show_error(self.tr("Error raised while saving it to Excel file.\n    {0}").format(e.__str__()))
                return
