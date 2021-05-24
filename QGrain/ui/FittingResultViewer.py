__all__ = ["FittingResultViewer"]

import copy
import logging
import pickle
import time
import typing
from collections import Counter
from uuid import UUID

import numpy as np
import qtawesome as qta
from PySide2.QtCore import QCoreApplication, Qt, Signal, QTimer
from PySide2.QtGui import QCursor
from PySide2.QtWidgets import (QAbstractItemView, QComboBox, QDialog,
                               QFileDialog, QGridLayout, QLabel, QMenu,
                               QMessageBox, QPushButton, QTableWidget,
                               QTableWidgetItem)
from QGrain.algorithms import DistributionType
from QGrain.algorithms.AsyncFittingWorker import AsyncFittingWorker
from QGrain.algorithms.distributions import get_distance_func_by_name
from QGrain.statistic import logarithmic
from QGrain.charts.BoxplotChart import BoxplotChart
from QGrain.charts.DistanceCurveChart import DistanceCurveChart
from QGrain.charts.MixedDistributionChart import MixedDistributionChart
from QGrain.charts.SSUTypicalComponentChart import SSUTypicalComponentChart
from QGrain.models.ClassicResolverSetting import built_in_distances
from QGrain.models.FittingResult import FittingResult
from QGrain.models.FittingTask import FittingTask
from QGrain.models.GrainSizeSample import GrainSizeSample
from QGrain.ui.ReferenceResultViewer import ReferenceResultViewer
from sklearn.cluster import KMeans
import openpyxl
from QGrain.use_excel import prepare_styles, column_to_char
from QGrain import QGRAIN_VERSION

class FittingResultViewer(QDialog):
    PAGE_ROWS = 20
    logger = logging.getLogger("root.QGrain.ui.FittingResultViewer")
    result_marked = Signal(FittingResult)
    def __init__(self, reference_viewer: ReferenceResultViewer, parent=None):
        super().__init__(parent=parent, f=Qt.Window)
        self.setWindowTitle(self.tr("SSU Fitting Result Viewer"))
        self.__fitting_results = [] # type: list[FittingResult]
        self.retry_tasks = {} # type: dict[UUID, FittingTask]
        self.__reference_viewer = reference_viewer
        self.init_ui()
        self.boxplot_chart = BoxplotChart(parent=self, toolbar=True)
        self.typical_chart = SSUTypicalComponentChart(parent=self, toolbar=True)
        self.distance_chart = DistanceCurveChart(parent=self, toolbar=True)
        self.mixed_distribution_chart = MixedDistributionChart(parent=self, toolbar=True, use_animation=True)
        self.file_dialog = QFileDialog(parent=self)
        self.async_worker = AsyncFittingWorker()
        self.async_worker.background_worker.task_succeeded.connect(self.on_fitting_succeeded)
        self.async_worker.background_worker.task_failed.connect(self.on_fitting_failed)
        self.update_page_list()
        self.update_page(self.page_index)
        self.msg_box = QMessageBox(self)
        self.msg_box.setWindowFlags(Qt.Drawer)

        self.outlier_msg_box = QMessageBox(self)
        self.outlier_msg_box.setWindowFlags(Qt.Drawer)
        self.outlier_msg_box.setStandardButtons(QMessageBox.Discard|QMessageBox.Retry|QMessageBox.Ignore)
        self.retry_progress_msg_box = QMessageBox()
        self.retry_progress_msg_box.setWindowFlags(Qt.CustomizeWindowHint|Qt.WindowTitleHint)
        self.retry_progress_msg_box.addButton(QMessageBox.Ok)
        self.retry_progress_msg_box.button(QMessageBox.Ok).hide()
        self.retry_progress_msg_box.setWindowTitle(self.tr("Progress"))
        self.retry_timer = QTimer(self)
        self.retry_timer.setSingleShot(True)
        self.retry_timer.timeout.connect(lambda: self.retry_progress_msg_box.exec_())

    def init_ui(self):
        self.data_table = QTableWidget(100, 100)
        self.data_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.data_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.data_table.setAlternatingRowColors(True)
        self.data_table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.main_layout = QGridLayout(self)
        self.main_layout.addWidget(self.data_table, 0, 0, 1, 3)

        self.previous_button = QPushButton(qta.icon("mdi.skip-previous-circle"), self.tr("Previous"))
        self.previous_button.setToolTip(self.tr("Click to back to the previous page."))
        self.previous_button.clicked.connect(self.on_previous_button_clicked)
        self.current_page_combo_box = QComboBox()
        self.current_page_combo_box.addItem(self.tr("Page {0}").format(1))
        self.current_page_combo_box.currentIndexChanged.connect(self.update_page)
        self.next_button = QPushButton(qta.icon("mdi.skip-next-circle"), self.tr("Next"))
        self.next_button.setToolTip(self.tr("Click to jump to the next page."))
        self.next_button.clicked.connect(self.on_next_button_clicked)
        self.main_layout.addWidget(self.previous_button, 1, 0)
        self.main_layout.addWidget(self.current_page_combo_box, 1, 1)
        self.main_layout.addWidget(self.next_button, 1, 2)

        self.distance_label = QLabel(self.tr("Distance"))
        self.distance_label.setToolTip(self.tr("It's the function to calculate the difference (on the contrary, similarity) between two samples."))
        self.distance_combo_box = QComboBox()
        self.distance_combo_box.addItems(built_in_distances)
        self.distance_combo_box.setCurrentText("log10MSE")
        self.distance_combo_box.currentTextChanged.connect(lambda: self.update_page(self.page_index))
        self.main_layout.addWidget(self.distance_label, 2, 0)
        self.main_layout.addWidget(self.distance_combo_box, 2, 1, 1, 2)
        self.menu = QMenu(self.data_table)
        self.menu.setShortcutAutoRepeat(True)

        self.mark_action = self.menu.addAction(qta.icon("mdi.marker-check"), self.tr("Mark Selection(s) as Reference"))
        self.mark_action.triggered.connect(self.mark_selections)
        self.remove_action = self.menu.addAction(qta.icon("fa.remove"), self.tr("Remove Selection(s)"))
        self.remove_action.triggered.connect(self.remove_selections)
        self.plot_loss_chart_action = self.menu.addAction(qta.icon("mdi.chart-timeline-variant"), self.tr("Plot Loss Chart"))
        self.plot_loss_chart_action.triggered.connect(self.show_distance)
        self.plot_distribution_chart_action = self.menu.addAction(qta.icon("fa5s.chart-area"), self.tr("Plot Distribution Chart"))
        self.plot_distribution_chart_action.triggered.connect(self.show_distribution)
        self.plot_distribution_animation_action = self.menu.addAction(qta.icon("fa5s.chart-area"), self.tr("Plot Distribution Chart (Animation)"))
        self.plot_distribution_animation_action.triggered.connect(self.show_history_distribution)
        self.detect_outliers_menu = self.menu.addMenu(qta.icon("mdi.magnify"), self.tr("Detect Outliers"))
        self.check_final_distances_action = self.detect_outliers_menu.addAction(self.tr("Check Final Distances"))
        self.check_final_distances_action.triggered.connect(self.check_final_distances)
        self.check_component_mean_action = self.detect_outliers_menu.addAction(self.tr("Check Component Mean"))
        self.check_component_mean_action.triggered.connect(lambda: self.check_component_moments("mean"))
        self.check_component_std_action = self.detect_outliers_menu.addAction(self.tr("Check Component STD"))
        self.check_component_std_action.triggered.connect(lambda: self.check_component_moments("std"))
        self.check_component_skewness_action = self.detect_outliers_menu.addAction(self.tr("Check Component Skewness"))
        self.check_component_skewness_action.triggered.connect(lambda: self.check_component_moments("skewness"))
        self.check_component_kurtosis_action = self.detect_outliers_menu.addAction(self.tr("Check Component Kurtosis"))
        self.check_component_kurtosis_action.triggered.connect(lambda: self.check_component_moments("kurtosis"))
        self.check_component_fractions_action = self.detect_outliers_menu.addAction(self.tr("Check Component Fractions"))
        self.check_component_fractions_action.triggered.connect(self.check_component_fractions)
        self.degrade_results_action = self.detect_outliers_menu.addAction(self.tr("Degrade Results"))
        self.degrade_results_action.triggered.connect(self.degrade_results)
        self.try_align_components_action = self.detect_outliers_menu.addAction(self.tr("Try Align Components"))
        self.try_align_components_action.triggered.connect(self.try_align_components)

        self.analyse_typical_components_action = self.menu.addAction(qta.icon("ei.tags"), self.tr("Analyse Typical Components"))
        self.analyse_typical_components_action.triggered.connect(self.analyse_typical_components)
        self.load_dump_action = self.menu.addAction(qta.icon("fa.database"), self.tr("Load Binary Dump"))
        self.load_dump_action.triggered.connect(self.load_dump)
        self.save_dump_action = self.menu.addAction(qta.icon("fa.save"), self.tr("Save Binary Dump"))
        self.save_dump_action.triggered.connect(self.save_dump)
        self.save_excel_action = self.menu.addAction(qta.icon("mdi.microsoft-excel"), self.tr("Save Excel"))
        self.save_excel_action.triggered.connect(lambda: self.on_save_excel_clicked(align_components=False))
        self.save_excel_align_action = self.menu.addAction(qta.icon("mdi.microsoft-excel"), self.tr("Save Excel (Force Alignment)"))
        self.save_excel_align_action.triggered.connect(lambda: self.on_save_excel_clicked(align_components=True))
        self.data_table.customContextMenuRequested.connect(self.show_menu)
        # necessary to add actions of menu to this widget itself,
        # otherwise, the shortcuts will not be triggered
        self.addActions(self.menu.actions())

    def show_menu(self, pos):
        self.menu.popup(QCursor.pos())

    def show_message(self, title: str, message: str):
        self.msg_box.setWindowTitle(title)
        self.msg_box.setText(message)
        self.msg_box.exec_()

    def show_info(self, message: str):
        self.show_message(self.tr("Info"), message)

    def show_warning(self, message: str):
        self.show_message(self.tr("Warning"), message)

    def show_error(self, message: str):
        self.show_message(self.tr("Error"), message)

    @property
    def distance_name(self) -> str:
        return self.distance_combo_box.currentText()

    @property
    def distance_func(self) -> typing.Callable:
        return get_distance_func_by_name(self.distance_combo_box.currentText())

    @property
    def page_index(self) -> int:
        return self.current_page_combo_box.currentIndex()

    @property
    def n_pages(self) -> int:
        return self.current_page_combo_box.count()

    @property
    def n_results(self) -> int:
        return len(self.__fitting_results)

    @property
    def selections(self):
        start = self.page_index*self.PAGE_ROWS
        temp = set()
        for item in self.data_table.selectedRanges():
            for i in range(item.topRow(), min(self.PAGE_ROWS+1, item.bottomRow()+1)):
                temp.add(i+start)
        indexes = list(temp)
        indexes.sort()
        return indexes

    def update_page_list(self):
        last_page_index = self.page_index
        if self.n_results == 0:
            n_pages = 1
        else:
            n_pages, left = divmod(self.n_results, self.PAGE_ROWS)
            if left != 0:
                n_pages += 1
        self.current_page_combo_box.blockSignals(True)
        self.current_page_combo_box.clear()
        self.current_page_combo_box.addItems([self.tr("Page {0}").format(i+1) for i in range(n_pages)])
        if last_page_index >= n_pages:
            self.current_page_combo_box.setCurrentIndex(n_pages-1)
        else:
            self.current_page_combo_box.setCurrentIndex(last_page_index)
        self.current_page_combo_box.blockSignals(False)

    def update_page(self, page_index: int):
        def write(row: int, col: int, value: str):
            if isinstance(value, str):
                pass
            elif isinstance(value, int):
                value = str(value)
            elif isinstance(value, float):
                value = f"{value: 0.4f}"
            else:
                value = value.__str__()
            item = QTableWidgetItem(value)
            item.setTextAlignment(Qt.AlignCenter)
            self.data_table.setItem(row, col, item)
        # necessary to clear
        self.data_table.clear()
        if page_index == self.n_pages - 1:
            start = page_index * self.PAGE_ROWS
            end = self.n_results
        else:
            start, end = page_index * self.PAGE_ROWS, (page_index+1) * self.PAGE_ROWS
        self.data_table.setRowCount(end-start)
        self.data_table.setColumnCount(7)
        self.data_table.setHorizontalHeaderLabels([
            self.tr("Resolver"),
            self.tr("Distribution Type"),
            self.tr("N_components"),
            self.tr("N_iterations"),
            self.tr("Spent Time [s]"),
            self.tr("Final Distance"),
            self.tr("Has Reference")])
        sample_names = [result.sample.name for result in self.__fitting_results[start: end]]
        self.data_table.setVerticalHeaderLabels(sample_names)
        for row, result in enumerate(self.__fitting_results[start: end]):
            write(row, 0, result.task.resolver)
            write(row, 1, self.get_distribution_name(result.task.distribution_type))
            write(row, 2, result.task.n_components)
            write(row, 3, result.n_iterations)
            write(row, 4, result.time_spent)
            write(row, 5, self.distance_func(result.sample.distribution, result.distribution))
            has_ref = result.task.initial_guess is not None or result.task.reference is not None
            write(row, 6, self.tr("Yes") if has_ref else self.tr("No"))

        self.data_table.resizeColumnsToContents()

    def on_previous_button_clicked(self):
        if self.page_index > 0:
            self.current_page_combo_box.setCurrentIndex(self.page_index-1)

    def on_next_button_clicked(self):
        if self.page_index < self.n_pages - 1:
            self.current_page_combo_box.setCurrentIndex(self.page_index+1)

    def get_distribution_name(self, distribution_type: DistributionType):
        if distribution_type == DistributionType.Normal:
            return self.tr("Normal")
        elif distribution_type == DistributionType.Weibull:
            return self.tr("Weibull")
        elif distribution_type == DistributionType.SkewNormal:
            return self.tr("Skew Normal")
        else:
            raise NotImplementedError(distribution_type)

    def add_result(self, result: FittingResult):
        if self.n_results == 0 or \
            (self.page_index == self.n_pages - 1 and \
            divmod(self.n_results, self.PAGE_ROWS)[-1] != 0):
            need_update = True
        else:
            need_update = False
        self.__fitting_results.append(result)
        self.update_page_list()
        if need_update:
            self.update_page(self.page_index)

    def add_results(self, results: typing.List[FittingResult]):
        if self.n_results == 0 or \
            (self.page_index == self.n_pages - 1 and \
            divmod(self.n_results, self.PAGE_ROWS)[-1] != 0):
            need_update = True
        else:
            need_update = False
        self.__fitting_results.extend(results)
        self.update_page_list()
        if need_update:
            self.update_page(self.page_index)

    def mark_selections(self):
        for index in self.selections:
            self.result_marked.emit(self.__fitting_results[index])

    def remove_results(self, indexes):
        results = []
        for i in reversed(indexes):
            res = self.__fitting_results.pop(i)
            results.append(res)
        self.update_page_list()
        self.update_page(self.page_index)

    def remove_selections(self):
        indexes = self.selections
        self.remove_results(indexes)

    def show_distance(self):
        results = [self.__fitting_results[i] for i in self.selections]
        if results is None or len(results) == 0:
            return
        result = results[0]
        self.distance_chart.show_distance_series(
            result.get_distance_series(self.distance_name),
            title=result.sample.name)
        self.distance_chart.show()

    def show_distribution(self):
        results = [self.__fitting_results[i] for i in self.selections]
        if results is None or len(results) == 0:
            return
        result = results[0]
        self.mixed_distribution_chart.show_model(result.view_model)
        self.mixed_distribution_chart.show()

    def show_history_distribution(self):
        results = [self.__fitting_results[i] for i in self.selections]
        if results is None or len(results) == 0:
            return
        result = results[0]
        self.mixed_distribution_chart.show_result(result)
        self.mixed_distribution_chart.show()

    def load_dump(self):
        filename, _ = self.file_dialog.getOpenFileName(
            self, self.tr("Select a binary dump file of SSU results"),
            None, self.tr("Binary dump (*.dump)"))
        if filename is None or filename == "":
            return
        with open(filename, "rb") as f:
            results = pickle.load(f) # type: list[FittingResult]
            valid = True
            if isinstance(results, list):
                for result in results:
                    if not isinstance(result, FittingResult):
                        valid = False
                        break
            else:
                valid = False

            if valid:
                if self.n_results != 0 and len(results) != 0:
                    old_classes = self.__fitting_results[0].classes_φ
                    new_classes = results[0].classes_φ
                    classes_inconsistent = False
                    if len(old_classes) != len(new_classes):
                        classes_inconsistent = True
                    else:
                        classes_error = np.abs(old_classes - new_classes)
                        if not np.all(np.less_equal(classes_error, 1e-8)):
                            classes_inconsistent = True
                    if classes_inconsistent:
                        self.show_error(self.tr("The results in the dump file has inconsistent grain-size classes with that in your list."))
                        return
                self.add_results(results)
            else:
                self.show_error(self.tr("The binary dump file is invalid."))

    def save_dump(self):
        if self.n_results == 0:
            self.show_warning(self.tr("There is not any result in the list."))
            return
        filename, _  = self.file_dialog.getSaveFileName(self, self.tr("Save the SSU results to binary dump file"),
                                            None, self.tr("Binary dump (*.dump)"))
        if filename is None or filename == "":
            return
        with open(filename, "wb") as f:
            pickle.dump(self.__fitting_results, f)

    def save_excel(self, filename, align_components=False):
        if self.n_results == 0:
            return

        results = self.__fitting_results.copy()
        classes_μm = results[0].classes_μm
        n_components_list = [result.n_components for result in self.__fitting_results]
        count_dict = Counter(n_components_list)
        max_n_components = max(count_dict.keys())
        self.logger.debug(f"N_components: {count_dict}, Max N_components: {max_n_components}")

        flags = []
        if not align_components:
            for result in results:
                flags.extend(range(result.n_components))
        else:
            n_components_desc = "\n".join([self.tr("{0} Component(s): {1}").format(n_components, count) for n_components, count in count_dict.items()])
            self.show_info(self.tr("N_components distribution of Results:\n{0}").format(n_components_desc))
            stacked_components = []
            for result in self.__fitting_results:
                for component in result.components:
                    stacked_components.append(component.distribution)
            stacked_components = np.array(stacked_components)
            cluser = KMeans(n_clusters=max_n_components)
            flags = cluser.fit_predict(stacked_components)
            # check flags to make it unique
            flag_index = 0
            for i, result in enumerate(self.__fitting_results):
                result_flags = set()
                for component in result.components:
                    if flags[flag_index] in result_flags:
                        if flags[flag_index] == max_n_components:
                            flags[flag_index] = max_n_components-1
                        else:
                            flag_index[flag_index] += 1
                        result_flags.add(flags[flag_index])
                    flag_index += 1

            flag_set = set(flags)
            picked = []
            for target_flag in flag_set:
                for i, flag in enumerate(flags):
                    if flag == target_flag:
                        picked.append((target_flag, logarithmic(classes_μm, stacked_components[i])["mean"]))
                        break
            picked.sort(key=lambda x: x[1])
            flag_map = {flag: index for index, (flag, _) in enumerate(picked)}
            flags = np.array([flag_map[flag] for flag in flags])

        wb = openpyxl.Workbook()
        prepare_styles(wb)
        ws = wb.active
        ws.title = self.tr("README")
        description = \
            """
            This Excel file was generated by QGrain ({0}).

            It contanins 4 + max(N_components) sheets:
            1. The first sheet is the sample distributions of SSU results.
            2. The second sheet is used to put the infomation of fitting.
            3. The third sheet is the statistic parameters calculated by statistic moment method.
            4. The fouth sheet is the distributions of unmixed components and their sum of each sample.
            5. Other sheets are the unmixed end-member distributions which were discretely stored.

            The SSU algorithm is implemented by QGrain.

            """.format(QGRAIN_VERSION)

        def write(row, col, value, style="normal_light"):
            cell = ws.cell(row+1, col+1, value=value)
            cell.style = style

        lines_of_desc = description.split("\n")
        for row, line in enumerate(lines_of_desc):
            write(row, 0, line, style="description")
        ws.column_dimensions[column_to_char(0)].width = 200

        ws = wb.create_sheet(self.tr("Sample Distributions"))
        write(0, 0, self.tr("Sample Name"), style="header")
        ws.column_dimensions[column_to_char(0)].width = 16
        for col, value in enumerate(classes_μm, 1):
            write(0, col, value, style="header")
            ws.column_dimensions[column_to_char(col)].width = 10
        for row, result in enumerate(results, 1):
            if row % 2 == 0:
                style = "normal_dark"
            else:
                style = "normal_light"
            write(row, 0, result.sample.name, style=style)
            for col, value in enumerate(result.sample.distribution, 1):
                write(row, col, value, style=style)
            QCoreApplication.processEvents()

        ws = wb.create_sheet(self.tr("Information of Fitting"))
        write(0, 0, self.tr("Sample Name"), style="header")
        ws.column_dimensions[column_to_char(0)].width = 16
        headers = [self.tr("Distribution Type"),
                   self.tr("N_components"),
                   self.tr("Resolver"),
                   self.tr("Resolver Settings"),
                   self.tr("Initial Guess"),
                   self.tr("Reference"),
                   self.tr("Spent Time [s]"),
                   self.tr("N_iterations"),
                   self.tr("Final Distance [log10MSE]")]
        for col, value in enumerate(headers, 1):
            write(0, col, value, style="header")
            if col in (4, 5, 6):
                ws.column_dimensions[column_to_char(col)].width = 10
            else:
                ws.column_dimensions[column_to_char(col)].width = 10
        for row, result in enumerate(results, 1):
            if row % 2 == 0:
                style = "normal_dark"
            else:
                style = "normal_light"
            write(row, 0, result.sample.name, style=style)
            write(row, 1, result.distribution_type.name, style=style)
            write(row, 2, result.n_components, style=style)
            write(row, 3, result.task.resolver, style=style)
            write(row, 4, self.tr("Default") if result.task.resolver_setting is None else result.task.resolver_setting.__str__(), style=style)
            write(row, 5, self.tr("None") if result.task.initial_guess is None else result.task.initial_guess.__str__(), style=style)
            write(row, 6, self.tr("None") if result.task.reference is None else result.task.reference.__str__(), style=style)
            write(row, 7, result.time_spent, style=style)
            write(row, 8, result.n_iterations, style=style)
            write(row, 9, result.get_distance("log10MSE"), style=style)

        ws = wb.create_sheet(self.tr("Statistic Moments"))
        write(0, 0, self.tr("Sample Name"), style="header")
        ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
        ws.column_dimensions[column_to_char(0)].width = 16
        headers = []
        sub_headers = [self.tr("Proportion"),
                       self.tr("Mean [φ]"),
                       self.tr("Mean [μm]"),
                       self.tr("STD [φ]"),
                       self.tr("STD [μm]"),
                       self.tr("Skewness"),
                       self.tr("Kurtosis")]
        for i in range(max_n_components):
            write(0, i*len(sub_headers)+1, self.tr("C{0}").format(i+1), style="header")
            ws.merge_cells(start_row=1, start_column=i*len(sub_headers)+2, end_row=1, end_column=(i+1)*len(sub_headers)+1)
            headers.extend(sub_headers)
        for col, value in enumerate(headers, 1):
            write(1, col, value, style="header")
            ws.column_dimensions[column_to_char(col)].width = 10
        flag_index = 0
        for row, result in enumerate(results, 2):
            if row % 2 == 0:
                style = "normal_light"
            else:
                style = "normal_dark"
            write(row, 0, result.sample.name, style=style)
            for component in result.components:
                index = flags[flag_index]
                write(row, index*len(sub_headers)+1, component.fraction, style=style)
                write(row, index*len(sub_headers)+2, component.logarithmic_moments["mean"], style=style)
                write(row, index*len(sub_headers)+3, component.geometric_moments["mean"], style=style)
                write(row, index*len(sub_headers)+4, component.logarithmic_moments["std"], style=style)
                write(row, index*len(sub_headers)+5, component.geometric_moments["std"], style=style)
                write(row, index*len(sub_headers)+6, component.logarithmic_moments["skewness"], style=style)
                write(row, index*len(sub_headers)+7, component.logarithmic_moments["kurtosis"], style=style)
                flag_index += 1

        ws = wb.create_sheet(self.tr("Unmixed Components"))
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
        write(0, 0, self.tr("Sample Name"), style="header")
        ws.column_dimensions[column_to_char(0)].width = 16
        for col, value in enumerate(classes_μm, 2):
            write(0, col, value, style="header")
            ws.column_dimensions[column_to_char(col)].width = 10
        row = 1
        for result_index, result in enumerate(results, 1):
            if result_index % 2 == 0:
                style = "normal_dark"
            else:
                style = "normal_light"
            write(row, 0, result.sample.name, style=style)
            ws.merge_cells(start_row=row+1, start_column=1, end_row=row+result.n_components+1, end_column=1)
            for component_i, component in enumerate(result.components, 1):
                write(row, 1, self.tr("C{0}").format(component_i), style=style)
                for col, value in enumerate(component.distribution*component.fraction, 2):
                    write(row, col, value, style=style)
                row += 1
            write(row, 1, self.tr("Sum"), style=style)
            for col, value in enumerate(result.distribution, 2):
                write(row, col, value, style=style)
            row += 1

        ws_dict = {}
        flag_set = set(flags)
        for flag in flag_set:
            ws = wb.create_sheet(self.tr("Unmixed EM{0}").format(flag+1))
            write(0, 0, self.tr("Sample Name"), style="header")
            ws.column_dimensions[column_to_char(0)].width = 16
            for col, value in enumerate(classes_μm, 1):
                write(0, col, value, style="header")
                ws.column_dimensions[column_to_char(col)].width = 10
            ws_dict[flag] = ws

        flag_index = 0
        for row, result in enumerate(results, 1):
            if row % 2 == 0:
                style = "normal_dark"
            else:
                style = "normal_light"

            for component in result.components:
                flag = flags[flag_index]
                ws = ws_dict[flag]
                write(row, 0, result.sample.name, style=style)
                for col, value in enumerate(component.distribution, 1):
                    write(row, col, value, style=style)
                flag_index += 1

        wb.save(filename)
        wb.close()

    def on_save_excel_clicked(self, align_components=False):
        if self.n_results == 0:
            self.show_warning(self.tr("There is not any SSU result."))
            return
        filename, _ = self.file_dialog.getSaveFileName(
            None, self.tr("Choose a filename to save SSU Results"),
            None, "Microsoft Excel (*.xlsx)")
        if filename is None or filename == "":
            return
        try:
            self.save_excel(filename, align_components)
            self.show_info(self.tr("SSU results have been saved to:\n    {0}").format(filename))
        except Exception as e:
            self.show_error(self.tr("Error raised while save SSU results to Excel file.\n    {0}").format(e.__str__()))

    def on_fitting_succeeded(self, result: FittingResult):
        result_replace_index = self.retry_tasks[result.task.uuid]
        self.__fitting_results[result_replace_index] = result
        self.retry_tasks.pop(result.task.uuid)
        self.retry_progress_msg_box.setText(self.tr("Tasks to be retried: {0}").format(len(self.retry_tasks)))
        if len(self.retry_tasks) == 0:
            self.retry_progress_msg_box.close()

        self.logger.debug(f"Retried task succeeded, sample name={result.task.sample.name}, distribution_type={result.task.distribution_type.name}, n_components={result.task.n_components}")
        self.update_page(self.page_index)

    def on_fitting_failed(self, failed_info: str, task: FittingTask):
        # necessary to remove it from the dict
        self.retry_tasks.pop(task.uuid)
        if len(self.retry_tasks) == 0:
            self.retry_progress_msg_box.close()
        self.show_error(self.tr("Failed to retry task, sample name={0}.\n{1}").format(task.sample.name, failed_info))
        self.logger.warning(f"Failed to retry task, sample name={task.sample.name}, distribution_type={task.distribution_type.name}, n_components={task.n_components}")

    def retry_results(self, indexes, results):
        assert len(indexes) == len(results)
        if len(results) == 0:
            return
        self.retry_progress_msg_box.setText(self.tr("Tasks to be retried: {0}").format(len(results)))
        self.retry_timer.start(1)
        for index, result in zip(indexes, results):
            query = self.__reference_viewer.query_reference(result.sample)
            ref_result = None
            if query is None:
                nearby_results = self.__fitting_results[index-5: index]+self.__fitting_results[index+1: index+6]
                ref_result = self.__reference_viewer.find_similar(result.sample, nearby_results)
            else:
                ref_result = query
            keys = ["mean", "std", "skewness"]
            reference = [{key: comp.logarithmic_moments[key] for key in keys} for comp in ref_result.components]
            task = FittingTask(result.sample,
                               ref_result.distribution_type,
                               ref_result.n_components,
                               resolver=ref_result.task.resolver,
                               resolver_setting=ref_result.task.resolver_setting,
                               reference=reference)
            self.logger.debug(f"Retry task: sample name={task.sample.name}, distribution_type={task.distribution_type.name}, n_components={task.n_components}")
            self.retry_tasks[task.uuid] = index
            self.async_worker.execute_task(task)

    def degrade_results(self):
        degrade_results = [] # type: list[FittingResult]
        degrade_indexes = [] # type: list[int]
        for i, result in enumerate(self.__fitting_results):
            for component in result.components:
                if component.fraction < 1e-3:
                    degrade_results.append(result)
                    degrade_indexes.append(i)
                    break
        self.logger.debug(f"Results should be degrade (have a redundant component): {[result.sample.name for result in degrade_results]}")
        if len(degrade_results) == 0:
            self.show_info(self.tr("No fitting result was evaluated as an outlier."))
            return
        self.show_info(self.tr("The results below should be degrade (have a redundant component:\n    {0}").format(
                ", ".join([result.sample.name for result in degrade_results])))

        self.retry_progress_msg_box.setText(self.tr("Tasks to be retried: {0}").format(len(degrade_results)))
        self.retry_timer.start(1)
        for index, result in zip(degrade_indexes, degrade_results):
            reference = []
            n_redundant = 0
            for component in result.components:
                if component.fraction < 1e-3:
                    n_redundant += 1
                else:
                    reference.append(dict(mean=component.logarithmic_moments["mean"],
                                          std=component.logarithmic_moments["std"],
                                          skewness=component.logarithmic_moments["skewness"]))
            task = FittingTask(result.sample,
                               result.distribution_type,
                               result.n_components-n_redundant if result.n_components > n_redundant else 1,
                               resolver=result.task.resolver,
                               resolver_setting=result.task.resolver_setting,
                               reference=reference)
            self.logger.debug(f"Retry task: sample name={task.sample.name}, distribution_type={task.distribution_type.name}, n_components={task.n_components}")
            self.retry_tasks[task.uuid] = index
            self.async_worker.execute_task(task)

    def ask_deal_outliers(self, outlier_results: typing.List[FittingResult],
                          outlier_indexes: typing.List[int]):
        assert len(outlier_indexes) == len(outlier_results)
        if len(outlier_results) == 0:
            self.show_info(self.tr("No fitting result was evaluated as an outlier."))
        else:
            self.outlier_msg_box.setText(self.tr("The fitting results have the component that its fraction is near zero:\n    {0}\nHow to deal with them?").format(
                    ", ".join([result.sample.name for result in outlier_results])))
            res = self.outlier_msg_box.exec_()
            if res == QMessageBox.Discard:
                self.remove_results(outlier_indexes)
            elif res == QMessageBox.Retry:
                self.retry_results(outlier_indexes, outlier_results)
            else:
                pass

    def check_final_distances(self):
        if self.n_results == 0:
            self.show_warning(self.tr("There is not any result in the list."))
            return
        elif self.n_results < 10:
            self.show_warning(self.tr("The results in list are too less."))
            return
        distances = []
        for result in self.__fitting_results:
            distances.append(result.get_distance(self.distance_name))
        distances = np.array(distances)
        self.boxplot_chart.show_dataset([distances], xlabels=[self.distance_name], ylabel=self.tr("Distance"))
        self.boxplot_chart.show()

        # calculate the 1/4, 1/2, and 3/4 postion value to judge which result is invalid
        # 1. the mean squared errors are much higher in the results which are lack of components
        # 2. with the component number getting higher, the mean squared error will get lower and finally reach the minimum
        median = np.median(distances)
        upper_group = distances[np.greater(distances, median)]
        lower_group = distances[np.less(distances, median)]
        value_1_4 = np.median(lower_group)
        value_3_4 = np.median(upper_group)
        distance_QR = value_3_4 - value_1_4
        outlier_results = []
        outlier_indexes = []
        for i, (result, distance) in enumerate(zip(self.__fitting_results, distances)):
            if distance > value_3_4 + distance_QR * 1.5:
            # which error too small is not outlier
            # if distance > value_3_4 + distance_QR * 1.5 or distance < value_1_4 - distance_QR * 1.5:
                outlier_results.append(result)
                outlier_indexes.append(i)
        self.logger.debug(f"Outlier results with too greater distances: {[result.sample.name for result in outlier_results]}")
        self.ask_deal_outliers(outlier_results, outlier_indexes)

    def check_component_moments(self, key: str):
        if self.n_results == 0:
            self.show_warning(self.tr("There is not any result in the list."))
            return
        elif self.n_results < 10:
            self.show_warning(self.tr("The results in list are too less."))
            return
        stacked_moments = []
        for result in self.__fitting_results:
            for component in result.components:
                stacked_moments.append(component.logarithmic_moments[key])
        stacked_moments = np.array(stacked_moments)

        key_trans = {"mean": self.tr("Mean"), "std": self.tr("STD"), "skewness": self.tr("Skewness"), "kurtosis": self.tr("Kurtosis")}
        key_label_trans = {"mean": self.tr("Mean [φ]"), "std": self.tr("STD [φ]"), "skewness": self.tr("Skewness"), "kurtosis": self.tr("Kurtosis")}
        self.boxplot_chart.show_dataset([stacked_moments], xlabels=[key_trans[key]], ylabel=key_label_trans[key])
        self.boxplot_chart.show()

        # calculate the 1/4, 1/2, and 3/4 postion value to judge which result is invalid
        # 1. the mean squared errors are much higher in the results which are lack of components
        # 2. with the component number getting higher, the mean squared error will get lower and finally reach the minimum
        median = np.median(stacked_moments)
        upper_group = stacked_moments[np.greater(stacked_moments, median)]
        lower_group = stacked_moments[np.less(stacked_moments, median)]
        value_1_4 = np.median(lower_group)
        value_3_4 = np.median(upper_group)
        distance_QR = value_3_4 - value_1_4
        outlier_results = []
        outlier_indexes = []
        for i, result in enumerate(self.__fitting_results):
            for component in result.components:
                distance = component.logarithmic_moments[key]
                if distance > value_3_4 + distance_QR * 1.5 or distance < value_1_4 - distance_QR * 1.5:
                    outlier_results.append(result)
                    outlier_indexes.append(i)
                    break
        self.logger.debug(f"Outlier results with abnormal {key} values of their components: {[result.sample.name for result in outlier_results]}")
        self.ask_deal_outliers(outlier_results, outlier_indexes)

    def check_component_fractions(self):
        outlier_results = []
        outlier_indexes = []
        for i, result in enumerate(self.__fitting_results):
            for component in result.components:
                if component.fraction < 1e-3:
                    outlier_results.append(result)
                    outlier_indexes.append(i)
                    break
        self.logger.debug(f"Outlier results with the component that its fraction is near zero: {[result.sample.name for result in outlier_results]}")
        self.ask_deal_outliers(outlier_results, outlier_indexes)

    def try_align_components(self):
        if self.n_results == 0:
            self.show_warning(self.tr("There is not any result in the list."))
            return
        elif self.n_results < 10:
            self.show_warning(self.tr("The results in list are too less."))
            return
        import matplotlib.pyplot as plt
        n_components_list = [result.n_components for result in self.__fitting_results]
        count_dict = Counter(n_components_list)
        max_n_components = max(count_dict.keys())
        self.logger.debug(f"N_components: {count_dict}, Max N_components: {max_n_components}")
        n_components_desc = "\n".join([self.tr("{0} Component(s): {1}").format(n_components, count) for n_components, count in count_dict.items()])
        self.show_info(self.tr("N_components distribution of Results:\n{0}").format(n_components_desc))

        x = self.__fitting_results[0].classes_μm
        stacked_components = []
        for result in self.__fitting_results:
            for component in result.components:
                stacked_components.append(component.distribution)
        stacked_components = np.array(stacked_components)

        cluser = KMeans(n_clusters=max_n_components)
        flags = cluser.fit_predict(stacked_components)

        figure = plt.figure(figsize=(6, 4))
        cmap = plt.get_cmap("tab10")
        axes = figure.add_subplot(1, 1, 1)
        for flag, distribution in zip(flags, stacked_components):
            plt.plot(x, distribution, c=cmap(flag), zorder=flag)
        axes.set_xscale("log")
        axes.set_xlabel(self.tr("Grain-size [μm]"))
        axes.set_ylabel(self.tr("Frequency"))
        figure.tight_layout()
        figure.show()

        outlier_results = []
        outlier_indexes = []
        flag_index = 0
        for i, result in enumerate(self.__fitting_results):
            result_flags = set()
            for component in result.components:
                if flags[flag_index] in result_flags:
                    outlier_results.append(result)
                    outlier_indexes.append(i)
                    break
                else:
                    result_flags.add(flags[flag_index])
                flag_index += 1
        self.logger.debug(f"Outlier results that have two components in the same cluster: {[result.sample.name for result in outlier_results]}")
        self.ask_deal_outliers(outlier_results, outlier_indexes)

    def analyse_typical_components(self):
        if self.n_results == 0:
            self.show_warning(self.tr("There is not any result in the list."))
            return
        elif self.n_results < 10:
            self.show_warning(self.tr("The results in list are too less."))
            return

        self.typical_chart.show_typical(self.__fitting_results)
        self.typical_chart.show()
