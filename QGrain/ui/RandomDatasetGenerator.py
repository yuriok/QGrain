import logging
import pickle
import typing

import numpy as np
import openpyxl
import qtawesome as qta
from PySide2.QtCore import QCoreApplication, QPointF, Qt, QTimer, Signal
from PySide2.QtWidgets import (QDialog, QDoubleSpinBox, QFileDialog,
                               QGridLayout, QGroupBox, QLabel, QProgressBar,
                               QPushButton, QSizePolicy, QSpinBox, QWidget)
from QGrain import QGRAIN_VERSION
from QGrain.charts.MixedDistributionChart import MixedDistributionChart
from QGrain.models.artificial import (LACUSTRINE, LOESS, ArtificialComponent,
                                      ArtificialDataset, ArtificialSample,
                                      RandomSetting, get_random_dataset)
from QGrain.use_excel import column_to_char, prepare_styles


class RandomGeneratorComponentWidget(QWidget):
    value_changed = Signal()
    PARAM_KEYS = ("shape", "loc", "scale", "weight")
    SHAPE_MEAN_RANGE = (-100.0, 100.0)
    SHAPE_MEAN_DEFAULT = 0.0
    SHAPE_MEAN_STEP = 0.1
    SHAPE_STD_RANGE = (0.0, 10.0)
    SHAPE_STD_DEFAULT = 0.0
    SHAPE_STD_STEP = 0.1
    LOC_MEAN_RANGE = (-12.0, 12.0)
    LOC_MEAN_DEFAULT = 0.0
    LOC_MEAN_STEP = 0.1
    LOC_STD_RANGE = (0.0, 10.0)
    LOC_STD_DEFAULT = 0.1
    LOC_STD_STEP = 0.1
    SCALE_MEAN_RANGE = (0.01, 100.0)  # CAN NOT EQUAL TO ZERO
    SCALE_MEAN_DEFAULT = 1.0
    SCALE_MEAN_STEP = 0.1
    SCALE_STD_RANGE = (0.0, 10.0)
    SCALE_STD_DEFAULT = 0.1
    SCALE_STD_STEP = 0.1
    WEIGHT_MEAN_RANGE = (1.0, 100.0)
    WEIGHT_MEAN_DEFAULT = 1.0
    WEIGHT_MEAN_STEP = 0.1
    WEIGHT_STD_RANGE = (0.0, 10.0)
    WEIGHT_STD_DEFAULT = 1.0
    WEIGHT_STD_STEP = 0.1

    def __init__(self, name: str, parent=None):
        super().__init__(parent=parent)
        self.setAttribute(Qt.WA_StyledBackground, True)
        self.init_ui(name)


    def init_ui(self, name: str):
        self.main_layout = QGridLayout(self)
        self.main_layout.setContentsMargins(0, 0, 0, 0)

        self.name_label = QLabel(name)
        self.name_label.setAlignment(Qt.AlignHCenter|Qt.AlignVCenter)
        self.name_label.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.MinimumExpanding)
        self.main_layout.addWidget(self.name_label, 1, 0, 4, 1)
        # shape
        self.shape_label = QLabel(self.tr("Shape"))
        self.shape_label.setAlignment(Qt.AlignHCenter|Qt.AlignVCenter)
        self.shape_mean_input = QDoubleSpinBox()
        self.shape_mean_input.setRange(*self.SHAPE_MEAN_RANGE)
        self.shape_mean_input.setSingleStep(self.SHAPE_MEAN_STEP)
        self.shape_mean_input.setValue(self.SHAPE_MEAN_DEFAULT)
        self.shape_std_input = QDoubleSpinBox()
        self.shape_std_input.setRange(*self.SHAPE_STD_RANGE)
        self.shape_std_input.setSingleStep(self.SHAPE_STD_STEP)
        self.shape_std_input.setValue(self.SHAPE_STD_DEFAULT)
        self.main_layout.addWidget(self.shape_label, 1, 1)
        self.main_layout.addWidget(self.shape_mean_input, 1, 2)
        self.main_layout.addWidget(self.shape_std_input, 1, 3)
        # loc
        self.loc_label = QLabel(self.tr("Loc"))
        self.loc_label.setAlignment(Qt.AlignHCenter|Qt.AlignVCenter)
        self.loc_mean_input = QDoubleSpinBox()
        self.loc_mean_input.setRange(*self.LOC_MEAN_RANGE)
        self.loc_mean_input.setSingleStep(self.LOC_MEAN_STEP)
        self.loc_mean_input.setValue(self.LOC_MEAN_DEFAULT)
        self.loc_std_input = QDoubleSpinBox()
        self.loc_std_input.setRange(*self.LOC_STD_RANGE)
        self.loc_std_input.setSingleStep(self.LOC_STD_STEP)
        self.loc_std_input.setValue(self.LOC_STD_DEFAULT)
        self.main_layout.addWidget(self.loc_label, 2, 1)
        self.main_layout.addWidget(self.loc_mean_input, 2, 2)
        self.main_layout.addWidget(self.loc_std_input, 2, 3)
        # scale
        self.scale_label = QLabel(self.tr("Scale"))
        self.scale_label.setAlignment(Qt.AlignHCenter|Qt.AlignVCenter)
        self.scale_mean_input = QDoubleSpinBox()
        self.scale_mean_input.setRange(*self.SCALE_MEAN_RANGE)
        self.scale_mean_input.setSingleStep(self.SCALE_MEAN_STEP)
        self.scale_mean_input.setValue(self.SCALE_MEAN_DEFAULT)
        self.scale_std_input = QDoubleSpinBox()
        self.scale_std_input.setRange(*self.SCALE_STD_RANGE)
        self.scale_std_input.setSingleStep(self.SCALE_STD_STEP)
        self.scale_std_input.setValue(self.SCALE_STD_DEFAULT)
        self.main_layout.addWidget(self.scale_label, 3, 1)
        self.main_layout.addWidget(self.scale_mean_input, 3, 2)
        self.main_layout.addWidget(self.scale_std_input, 3, 3)
        # weight
        self.weight_label = QLabel(self.tr("Weight"))
        self.weight_label.setAlignment(Qt.AlignHCenter|Qt.AlignVCenter)
        self.weight_mean_input = QDoubleSpinBox()
        self.weight_mean_input.setRange(*self.WEIGHT_MEAN_RANGE)
        self.weight_mean_input.setSingleStep(self.WEIGHT_MEAN_STEP)
        self.weight_mean_input.setValue(self.WEIGHT_MEAN_DEFAULT)
        self.weight_std_input = QDoubleSpinBox()
        self.weight_std_input.setRange(*self.WEIGHT_STD_RANGE)
        self.weight_std_input.setSingleStep(self.WEIGHT_STD_STEP)
        self.weight_std_input.setValue(self.WEIGHT_STD_DEFAULT)
        self.main_layout.addWidget(self.weight_label, 4, 1)
        self.main_layout.addWidget(self.weight_mean_input, 4, 2)
        self.main_layout.addWidget(self.weight_std_input, 4, 3)

        self.shape_mean_input.valueChanged.connect(self.on_value_changed)
        self.shape_std_input.valueChanged.connect(self.on_value_changed)
        self.loc_mean_input.valueChanged.connect(self.on_value_changed)
        self.loc_std_input.valueChanged.connect(self.on_value_changed)
        self.scale_mean_input.valueChanged.connect(self.on_value_changed)
        self.scale_std_input.valueChanged.connect(self.on_value_changed)
        self.weight_mean_input.valueChanged.connect(self.on_value_changed)
        self.weight_std_input.valueChanged.connect(self.on_value_changed)

    def get_param(self, key):
        if hasattr(self, f"{key}_input"):
            return getattr(self, f"{key}_input").value()
        else:
            raise KeyError(key)

    def set_param(self, key, value):
        if hasattr(self, f"{key}_input"):
            getattr(self, f"{key}_input").setValue(value)
        else:
            raise KeyError(key)

    @property
    def target(self):
        target = {key: (self.get_param(key+"_mean"), self.get_param(key+"_std")) for key in self.PARAM_KEYS}
        return target

    @target.setter
    def target(self, values: dict):
        for key, (mean, std) in values.items():
            self.set_param(key+"_mean", mean)
            self.set_param(key+"_std", std)

    def on_value_changed(self, _):
        self.value_changed.emit()

class RandomDatasetGenerator(QDialog):
    logger = logging.getLogger("root.ui.RandomGeneratorWidget")
    gui_logger = logging.getLogger("GUI")
    def __init__(self, parent=None):
        super().__init__(parent=parent, f=Qt.Window)
        self.setWindowTitle(self.tr("Random Dataset Generator"))
        self.last_n_components = 0
        self.components = [] # typing.List[RandomGeneratorComponentWidget]
        self.component_series = []
        self.init_ui()
        self.target = LOESS
        self.minimum_size_input.setValue(0.02)
        self.maximum_size_input.setValue(2000.0)
        self.n_classes_input.setValue(101)
        self.precision_input.setValue(4)

        self.file_dialog = QFileDialog(parent=self)
        self.update_timer = QTimer()
        self.update_timer.timeout.connect(lambda: self.update_chart(True))
        self.cancel_flag = False

    def init_ui(self):
        self.setAttribute(Qt.WA_StyledBackground, True)
        self.main_layout = QGridLayout(self)
        # self.main_layout.setContentsMargins(0, 0, 0, 0)

        self.sampling_group = QGroupBox(self.tr("Sampling"))
        # self.control_group.setFixedSize(400, 160)
        self.control_layout = QGridLayout(self.sampling_group)
        self.minimum_size_label = QLabel(self.tr("Minimum Size (μm)"))
        self.minimum_size_input = QDoubleSpinBox()
        self.minimum_size_input.setDecimals(2)
        self.minimum_size_input.setRange(1e-4, 1e6)
        self.minimum_size_input.setValue(0.0200)
        self.maximum_size_label = QLabel(self.tr("Maximum Size (μm)"))
        self.maximum_size_input = QDoubleSpinBox()
        self.maximum_size_input.setDecimals(2)
        self.maximum_size_input.setRange(1e-4, 1e6)
        self.maximum_size_input.setValue(2000.0000)
        self.control_layout.addWidget(self.minimum_size_label, 0, 0)
        self.control_layout.addWidget(self.minimum_size_input, 0, 1)
        self.control_layout.addWidget(self.maximum_size_label, 0, 2)
        self.control_layout.addWidget(self.maximum_size_input, 0, 3)
        self.n_classes_label = QLabel(self.tr("N Classes"))
        self.n_classes_input = QSpinBox()
        self.n_classes_input.setRange(10, 1e4)
        self.n_classes_input.setValue(101)
        self.precision_label = QLabel(self.tr("Data Precision"))
        self.precision_input = QSpinBox()
        self.precision_input.setRange(2, 8)
        self.precision_input.setValue(4)
        self.control_layout.addWidget(self.n_classes_label, 1, 0)
        self.control_layout.addWidget(self.n_classes_input, 1, 1)
        self.control_layout.addWidget(self.precision_label, 1, 2)
        self.control_layout.addWidget(self.precision_input, 1, 3)
        self.component_number_label = QLabel(self.tr("N Components"))
        self.component_number_input = QSpinBox()
        self.component_number_input.setRange(1, 10)
        self.component_number_input.valueChanged.connect(self.on_n_components_changed)
        self.preview_button = QPushButton(qta.icon("mdi.animation-play"), self.tr("Preview"))
        self.preview_button.clicked.connect(self.on_preview_clicked)
        self.control_layout.addWidget(self.component_number_label, 2, 0)
        self.control_layout.addWidget(self.component_number_input, 2, 1)
        self.control_layout.addWidget(self.preview_button, 2, 2, 1, 2)
        self.main_layout.addWidget(self.sampling_group, 0, 0)

        self.save_group = QGroupBox(self.tr("Save"))
        # self.save_group.setFixedHeight(160)
        self.save_layout = QGridLayout(self.save_group)
        self.n_samples_label = QLabel(self.tr("Sample Number"))
        self.n_samples_input = QSpinBox()
        self.n_samples_input.setRange(100, 100000)
        self.save_layout.addWidget(self.n_samples_label, 0, 0)
        self.save_layout.addWidget(self.n_samples_input, 0, 1)

        self.cancel_button = QPushButton(qta.icon("mdi.cancel"), self.tr("Cancel"))
        self.cancel_button.setEnabled(False)
        self.cancel_button.clicked.connect(self.on_cancel_clicked)
        self.generate_button = QPushButton(qta.icon("mdi.cube-send"), self.tr("Generate"))
        self.generate_button.clicked.connect(self.on_generate_clicked)
        self.progress_bar = QProgressBar()
        self.progress_bar.setRange(0, 100)
        self.progress_bar.setOrientation(Qt.Horizontal)
        self.progress_bar.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        self.save_layout.addWidget(self.cancel_button, 1, 0)
        self.save_layout.addWidget(self.generate_button, 1, 1)
        self.save_layout.addWidget(self.progress_bar, 2, 0, 1, 2)
        self.main_layout.addWidget(self.save_group, 0, 1)

        self.param_group = QGroupBox("Random Parameter")
        # self.param_group.setFixedWidth(400)
        self.param_layout = QGridLayout(self.param_group)
        self.main_layout.addWidget(self.param_group, 1, 0)

        self.preview_group = QGroupBox(self.tr("Preview"))
        self.chart_layout = QGridLayout(self.preview_group)

        self.chart = MixedDistributionChart(parent=self, toolbar=False)
        self.chart_layout.addWidget(self.chart, 0, 0)
        self.chart.setSizePolicy(QSizePolicy.MinimumExpanding, QSizePolicy.MinimumExpanding)
        self.main_layout.addWidget(self.preview_group, 1, 1)

    @staticmethod
    def to_points(x, y):
        return [QPointF(x_value, y_value) for x_value, y_value in zip(x, y)]

    def on_n_components_changed(self, n_components: int):
        if self.last_n_components < n_components:
            for component_index in range(self.last_n_components, n_components):
                component = RandomGeneratorComponentWidget(name=f"AC{component_index+1}")
                component.value_changed.connect(self.on_value_changed)
                self.param_layout.addWidget(component, component_index+1, 0)
                self.components.append(component)

        if self.last_n_components > n_components:
            for i in range(n_components, self.last_n_components):
                before_component = self.components[i]
                before_component.value_changed.disconnect(self.on_value_changed)
                self.param_layout.removeWidget(before_component)
                # if not hide, the widget will still display on screen
                before_component.hide()
                self.components.pop(n_components)

        self.last_n_components = n_components

    def on_preview_clicked(self):
        if self.update_timer.isActive():
            self.preview_button.setText(self.tr("Preview"))
            self.update_timer.stop()
            self.update_chart()
        else:
            self.preview_button.setText(self.tr("Stop"))
            self.update_timer.start(200)

    def on_cancel_clicked(self):
        self.cancel_flag = True

    def on_generate_clicked(self):
        if self.update_timer.isActive():
            self.preview_button.setText(self.tr("Preview"))
            self.update_timer.stop()
            self.update_chart()

        filename, _ = self.file_dialog.getSaveFileName(
            self, self.tr("Choose a filename to save the generated dataset"),
            None, "Microsoft Excel (*.xlsx)")
        if filename is None or filename == "":
            return
        n_samples = self.n_samples_input.value()
        dataset = self.get_random_dataset(n_samples)
        # generate samples
        self.cancel_button.setEnabled(True)
        self.generate_button.setEnabled(False)
        format_str = self.tr("Generating {0} samples: %p%").format(n_samples)
        self.progress_bar.setFormat(format_str)
        self.progress_bar.setValue(0)

        def cancel():
            self.progress_bar.setFormat(self.tr("Task canceled"))
            self.progress_bar.setValue(0)
            self.cancel_button.setEnabled(False)
            self.generate_button.setEnabled(True)
            self.cancel_flag = False

        samples = []
        for i in range(n_samples):
            if self.cancel_flag:
                cancel()
                return
            sample = dataset.get_sample(i)
            samples.append(sample)
            progress = (i+1) / n_samples * 50
            self.progress_bar.setValue(progress)
            QCoreApplication.processEvents()

        # save file to excel file
        format_str = self.tr("Writing {0} samples to excel file: %p%").format(n_samples)
        self.progress_bar.setFormat(format_str)
        self.progress_bar.setValue(50)

        wb = openpyxl.Workbook()
        prepare_styles(wb)

        ws = wb.active
        ws.title = self.tr("README")
        description = \
            """
            This Excel file was generated by QGrain ({0}).

            It contanins n_components + 3 sheets:
            1. The first sheet is the random settings which were used to generate random parameters.
            2. The second sheet is the generated dataset.
            3. The third sheet is random parameters which were used to calulate the component distributions and their mixture.
            4. The left sheets are the component distributions of all samples.

            Artificial dataset
                Using skew normal distribution as the base distribution of each component (i.e. end-member).
                Skew normal distribution has three parameters, shape, loc and scale.
                Where shape controls the skewness, loc and scale are simliar to that of the Normal distribution.
                When shape = 0, it becomes Normal distribution.
                The weight parameter controls the fraction of the component, where fraction_i = weight_i / sum(weight_i).
                By assigning the mean and std of each parameter, random parameters was generate by the `scipy.stats.truncnorm.rvs` function of Scipy.

            Sampling settings
                Minimum size [μm]: {1},
                Maximum size [μm]: {2},
                N_classes: {3},
                Precision: {4},
                Noise: {5},
                N_samples: {6},

            """.format(QGRAIN_VERSION,
                       self.minimum_size_input.value(),
                       self.maximum_size_input.value(),
                       self.n_classes_input.value(),
                       self.precision_input.value(),
                       self.precision_input.value()+1,
                       n_samples)

        def write(row, col, value, style="normal_light"):
            cell = ws.cell(row+1, col+1, value=value)
            cell.style = style

        lines_of_desc = description.split("\n")
        for row, line in enumerate(lines_of_desc):
            write(row, 0, line, style="description")
        ws.column_dimensions[column_to_char(0)].width = 200

        ws = wb.create_sheet(self.tr("Random settings"))
        write(0, 0, self.tr("Parameter"), style="header")
        ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
        write(0, 1, self.tr("Shape"), style="header")
        ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=3)
        write(0, 3, self.tr("Loc"), style="header")
        ws.merge_cells(start_row=1, start_column=4, end_row=1, end_column=5)
        write(0, 5, self.tr("Scale"), style="header")
        ws.merge_cells(start_row=1, start_column=6, end_row=1, end_column=7)
        write(0, 7, self.tr("Weight"), style="header")
        ws.merge_cells(start_row=1, start_column=8, end_row=1, end_column=9)
        ws.column_dimensions[column_to_char(0)].width = 16
        for col in range(1, 9):
            ws.column_dimensions[column_to_char(col)].width = 16
            if col % 2 == 0:
                write(1, col, self.tr("Mean"), style="header")
            else:
                write(1, col, self.tr("STD"), style="header")
        for row, comp_params in enumerate(self.target, 2):
            if row % 2 == 1:
                style = "normal_dark"
            else:
                style = "normal_light"
            write(row, 0, self.tr("Component{0}").format(row-1), style=style)
            for i, key in enumerate(["shape", "loc", "scale", "weight"]):
                mean, std = comp_params[key]
                write(row, i*2+1, mean, style=style)
                write(row, i*2+2, std, style=style)

        ws = wb.create_sheet(self.tr("Dataset"))
        write(0, 0, self.tr("Sample Name"), style="header")
        ws.column_dimensions[column_to_char(0)].width = 16
        for col, value in enumerate(dataset.classes_μm, 1):
            write(0, col, value, style="header")
            ws.column_dimensions[column_to_char(col)].width = 10
        for row, sample in enumerate(samples, 1):
            if row % 2 == 0:
                style = "normal_dark"
            else:
                style = "normal_light"
            write(row, 0, sample.name, style=style)
            for col, value in enumerate(sample.distribution, 1):
                write(row, col, value, style=style)

            if self.cancel_flag:
                cancel()
                return
            progress = 50 + (row / n_samples) * 10
            self.progress_bar.setValue(progress)
            QCoreApplication.processEvents()

        ws = wb.create_sheet(self.tr("Parameters"))
        write(0, 0, self.tr("Sample Name"), style="header")
        ws.column_dimensions[column_to_char(0)].width = 16
        for i in range(dataset.n_components):
            write(0, 4*i+1, self.tr("Component{0}").format(i+1), style="header")
            ws.merge_cells(start_row=1, start_column=4*i+2, end_row=1, end_column=4*i+5)
            for j, header_name in enumerate([self.tr("Shape"), self.tr("Loc"), self.tr("Scale"), self.tr("Weight")]):
                write(1, 4*i+1+j, header_name, style="header")
                ws.column_dimensions[column_to_char(4*i+1+j)].width = 16
        for row, sample in enumerate(samples, 1):
            if row % 2 == 0:
                style = "normal_dark"
            else:
                style = "normal_light"
            write(row, 0, sample.name, style=style)
            for i, comp_param in enumerate(sample.parameter.components):
                write(row, 4*i+1, comp_param.shape, style=style)
                write(row, 4*i+2, comp_param.loc, style=style)
                write(row, 4*i+3, comp_param.scale, style=style)
                write(row, 4*i+4, comp_param.weight, style=style)
            if self.cancel_flag:
                cancel()
                return
            progress = 60 + (row / n_samples) * 10
            self.progress_bar.setValue(progress)
            QCoreApplication.processEvents()

        for i in range(dataset.n_components):
            ws = wb.create_sheet(self.tr("Component{0}").format(i+1))
            write(0, 0, self.tr("Sample Name"), style="header")
            ws.column_dimensions[column_to_char(0)].width = 16
            for col, value in enumerate(dataset.classes_μm, 1):
                write(0, col, value, style="header")
                ws.column_dimensions[column_to_char(col)].width = 10
            for row, sample in enumerate(samples, 1):
                if row % 2 == 0:
                    style = "normal_dark"
                else:
                    style = "normal_light"
                write(row, 0, sample.name, style=style)
                for col, value in enumerate(sample.components[i].distribution, 1):
                    write(row, col, value, style=style)
            if self.cancel_flag:
                cancel()
                return
            progress = 70 + ((i*n_samples + row) / n_samples*dataset.n_components) * 30
            self.progress_bar.setValue(progress)
            QCoreApplication.processEvents()
        wb.save(filename)
        wb.close()

        self.progress_bar.setValue(100)
        self.progress_bar.setFormat(self.tr("Task finished"))
        self.cancel_button.setEnabled(False)
        self.generate_button.setEnabled(True)

    @property
    def target(self):
        return [comp.target for comp in self.components]

    @target.setter
    def target(self, values):
        if len(values) != len(self.components):
            self.component_number_input.setValue(len(values))
        for comp, comp_target in zip(self.components, values):
            comp.blockSignals(True)
            comp.target = comp_target
            comp.blockSignals(False)
        self.update_chart()

    def get_random_sample(self):
        dataset = self.get_random_dataset(n_samples=1)
        sample = dataset.get_sample(0)
        sample.name = self.tr("Artificial Sample")
        return sample

    def get_random_mean(self):
        dataset = self.get_random_dataset(n_samples=1)
        random_setting = RandomSetting(self.target)
        sample = dataset.get_sample_by_params(self.tr("Artificial Sample"), random_setting.mean_param)
        return sample

    def get_random_dataset(self, n_samples):
        min_μm = self.minimum_size_input.value()
        max_μm = self.maximum_size_input.value()
        n_classes = self.n_classes_input.value()
        if min_μm == max_μm:
            return
        if min_μm > max_μm:
            min_μm, max_μm = max_μm, min_μm
        precision = self.precision_input.value()
        noise = precision + 1

        dataset = get_random_dataset(target=self.target, n_samples=n_samples,
                                     min_μm=min_μm, max_μm=max_μm, n_classes=n_classes,
                                     precision=precision, noise=noise)
        return dataset

    def on_value_changed(self):
        self.update_chart()

    def update_chart(self, random=False):
        if not random:
            sample = self.get_random_mean()
        else:
            sample = self.get_random_sample()
        self.chart.show_model(sample.view_model)

    def closeEvent(self, event):
        if self.cancel_button.isEnabled():
            self.on_cancel_clicked()
        event.accept()


if __name__ == "__main__":
    import sys

    from QGrain.entry import setup_app
    app, splash = setup_app()
    main = RandomDatasetGenerator()
    main.show()
    splash.finish(main)
    sys.exit(app.exec_())
