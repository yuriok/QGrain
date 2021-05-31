import typing

import matplotlib.pyplot as plt
import numpy as np
import openpyxl
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.backends.backend_qt5agg import NavigationToolbar2QT as NavigationToolbar
from PySide2.QtCore import QCoreApplication, Qt
from PySide2.QtWidgets import (QComboBox, QDialog, QDoubleSpinBox, QFileDialog, QGridLayout,
                               QLabel, QMessageBox, QPushButton, QSpinBox)
from QGrain import QGRAIN_VERSION
from QGrain.statistic import convert_φ_to_μm
from QGrain.ssu import SSUResult
from QGrain.models.GrainSizeSample import GrainSizeSample
from QGrain.use_excel import column_to_char, prepare_styles
from sklearn.cluster import OPTICS


class SSUTypicalComponentChart(QDialog):
    def __init__(self, parent=None, toolbar=False):
        flags = Qt.Window | Qt.WindowTitleHint | Qt.CustomizeWindowHint | Qt.WindowCloseButtonHint
        super().__init__(parent=parent, f=flags)
        self.setWindowTitle(self.tr("SSU Typical Component Chart"))
        self.figure = plt.figure(figsize=(6, 3))
        self.clustering_axes = self.figure.add_subplot(1, 2, 1)
        self.component_axes = self.figure.add_subplot(1, 2, 2)
        self.canvas = FigureCanvas(self.figure)
        self.toolbar = NavigationToolbar(self.canvas, self)
        self.main_layout = QGridLayout(self)
        self.main_layout.addWidget(self.toolbar, 0, 0, 1, 4)
        self.main_layout.addWidget(self.canvas, 1, 0, 1, 4)
        if not toolbar:
            self.toolbar.hide()
        self.supported_scales = [("log-linear", self.tr("Log-linear")),
                                 ("log", self.tr("Log")),
                                 ("phi", self.tr("φ")),
                                 ("linear", self.tr("Linear"))]
        self.AXIS_LIST = [self.tr("Mean [φ]"), self.tr("Standard deviation [φ]"),
                          self.tr("Skewness"), self.tr("Kurtosis")]
        self.x_axis_label = QLabel(self.tr("X Axis"))
        self.x_axis_combo_box = QComboBox()
        self.x_axis_combo_box.addItems(self.AXIS_LIST)
        self.y_axis_label = QLabel(self.tr("Y Axis"))
        self.y_axis_combo_box = QComboBox()
        self.y_axis_combo_box.addItems(self.AXIS_LIST)
        self.y_axis_combo_box.setCurrentIndex(1)
        self.main_layout.addWidget(self.x_axis_label, 2, 0)
        self.main_layout.addWidget(self.x_axis_combo_box, 2, 1)
        self.main_layout.addWidget(self.y_axis_label, 3, 0)
        self.main_layout.addWidget(self.y_axis_combo_box, 3, 1)
        self.scale_label = QLabel(self.tr("Scale"))
        self.scale_combo_box = QComboBox()
        self.scale_combo_box.addItems([name for key, name in self.supported_scales])
        self.main_layout.addWidget(self.scale_label, 2, 2)
        self.main_layout.addWidget(self.scale_combo_box, 2, 3)
        self.min_samples_label = QLabel(self.tr("Minimum Samples"))
        self.min_samples_input = QDoubleSpinBox()
        self.min_samples_input.setRange(0.01, 0.99)
        self.min_samples_input.setValue(0.03)
        self.min_cluster_size_label = QLabel(self.tr("Minimum Cluster Size"))
        self.min_cluster_size_input = QDoubleSpinBox()
        self.min_cluster_size_input.setRange(0.01, 0.99)
        self.min_cluster_size_input.setValue(0.1)
        self.xi_label = QLabel(self.tr("xi"))
        self.xi_input = QDoubleSpinBox()
        self.xi_input.setRange(0.01, 0.99)
        self.xi_input.setValue(0.05)
        self.main_layout.addWidget(self.min_samples_label, 3, 2)
        self.main_layout.addWidget(self.min_samples_input, 3, 3)
        self.main_layout.addWidget(self.min_cluster_size_label, 4, 0)
        self.main_layout.addWidget(self.min_cluster_size_input, 4, 1)
        self.main_layout.addWidget(self.xi_label, 4, 2)
        self.main_layout.addWidget(self.xi_input, 4, 3)
        self.update_chart_button = QPushButton(self.tr("Update Chart"))
        self.update_chart_button.clicked.connect(self.update_chart)
        self.save_typical_button = QPushButton(self.tr("Save Typical"))
        self.save_typical_button.setEnabled(False)
        self.save_typical_button.clicked.connect(self.on_save_clicked)
        self.main_layout.addWidget(self.update_chart_button, 5, 0, 1, 2)
        self.main_layout.addWidget(self.save_typical_button, 5, 2, 1, 2)

        self.last_results = None # type: list[SSUResult]
        self.data_to_clustering = None
        self.stacked_components = None
        self.msg_box = QMessageBox(self)
        self.msg_box.setWindowFlags(Qt.Drawer)
        self.file_dialog = QFileDialog(parent=self)

    @property
    def scale(self) -> str:
        index = self.scale_combo_box.currentIndex()
        key, name = self.supported_scales[index]
        return key

    @property
    def transfer(self) -> typing.Callable:
        if self.scale == "log-linear":
            return lambda classes_φ: convert_φ_to_μm(classes_φ)
        elif self.scale == "log":
            return lambda classes_φ: np.log(convert_φ_to_μm(classes_φ))
        elif self.scale == "phi":
            return lambda classes_φ: classes_φ
        elif self.scale == "linear":
            return lambda classes_φ: convert_φ_to_μm(classes_φ)

    @property
    def xlabel(self) -> str:
        if self.scale == "log-linear":
            return self.tr("Grain-size [μm]")
        elif self.scale == "log":
            return self.tr("Ln(grain-size in μm)")
        elif self.scale == "phi":
            return self.tr("Grain-size [φ]")
        elif self.scale == "linear":
            return self.tr("Grain-size [μm]")

    @property
    def ylabel(self) -> str:
        return self.tr("Frequency")

    @property
    def xlog(self) -> bool:
        if self.scale == "log-linear":
            return True
        else:
            return False

    def show_message(self, title: str, message: str):
        self.msg_box.setWindowTitle(title)
        self.msg_box.setText(message)
        self.msg_box.exec_()

    def show_info(self, message: str):
        self.show_message(self.tr("Info"), message)

    def show_warning(self, message: str):
        self.show_message(self.tr("Warning"), message)

    def show_error(self, message: str):
        self.show_message(self.tr("Error"), message)

    def update_chart(self):
        if self.last_results is None:
            return
        x = self.transfer(self.last_results[0].classes_φ)
        self.save_typical_button.setEnabled(True)
        cluster = OPTICS(min_samples=self.min_samples_input.value(),
                         min_cluster_size=self.min_cluster_size_input.value(),
                         xi=self.xi_input.value())
        flags = cluster.fit_predict(self.data_to_clustering)
        cmap = plt.get_cmap()

        self.clustering_axes.clear()
        flag_set = set(flags)
        for flag in flag_set:
            key = np.equal(flags, flag)
            if flag == -1:
                c = "#7a7374"
                label = self.tr("Not clustered")
            else:
                c = cmap(flag)
                label = self.tr("EM{0}").format(flag+1)
            self.clustering_axes.plot(self.data_to_clustering[key][:, self.x_axis_combo_box.currentIndex()],
                                      self.data_to_clustering[key][:, self.y_axis_combo_box.currentIndex()],
                                      c="#ffffff00", marker=".", ms=8, mfc=c, mew=0.0,
                                      zorder=flag,
                                      label=label)
        if len(flag_set) < 6:
            self.clustering_axes.legend(loc="upper left")
        self.clustering_axes.set_xlabel(self.x_axis_combo_box.currentText())
        self.clustering_axes.set_ylabel(self.y_axis_combo_box.currentText())
        self.clustering_axes.set_title(self.tr("Clustering of end-members"))

        self.component_axes.clear()
        if self.xlog:
            self.component_axes.set_xscale("log")

        for flag in flag_set:
            if flag == -1:
                c = "#7a7374"
            else:
                c = cmap(flag)
            key = np.equal(flags, flag)
            for distribution in self.stacked_components[key]:
                self.component_axes.plot(x, distribution, c=c, zorder=flag)

            if flag != -1:
                typical = np.mean(self.stacked_components[key], axis=0)
                self.component_axes.plot(x, typical, c="black", zorder=1e10, ls="--", linewidth=1)
        self.component_axes.set_title(self.tr("Typical end-members"))
        self.component_axes.set_xlabel(self.xlabel)
        self.component_axes.set_ylabel(self.ylabel)
        self.component_axes.set_xlim(x[0], x[-1])
        self.component_axes.set_ylim(0, None)

        self.figure.tight_layout()
        self.canvas.draw()

    def show_typical(self, results: typing.Iterable[GrainSizeSample]):
        if len(results) == 0:
            return
        keys_to_clustering = ["mean", "std", "skewness", "kurtosis"]
        data_to_clustering = []
        stacked_components = []
        for result in results:
            for component in result.components:
                data_to_clustering.append([component.logarithmic_moments[key] for key in keys_to_clustering])
                stacked_components.append(component.distribution)
        # convert to numpy array
        data_to_clustering = np.array(data_to_clustering)
        stacked_components = np.array(stacked_components)
        self.last_results = results
        self.data_to_clustering = data_to_clustering
        self.stacked_components = stacked_components
        self.update_chart()

    def save_typical(self, filename):
        assert self.last_results is not None
        if len(self.last_results) == 0:
            return
        cluster = OPTICS(min_samples=self.min_samples_input.value(),
                         min_cluster_size=self.min_cluster_size_input.value(),
                         xi=self.xi_input.value())
        classes_μm = self.last_results[0].classes_μm
        flags = cluster.fit_predict(self.data_to_clustering)
        flag_set = set(flags)
        typicals = []
        for flag in flag_set:
            if flag != -1:
                key = np.equal(flags, flag)
                typical = np.mean(self.stacked_components[key], axis=0)
                typicals.append(typical)

        wb = openpyxl.Workbook()
        prepare_styles(wb)
        ws = wb.active
        ws.title = self.tr("README")
        description = \
            """
            This Excel file was generated by QGrain ({0}).

            It contanins 2 + N_clusters sheets:
            1. The first sheet is the sum distributions of all component clusters.
            2. The second sheet is used to put the component distributions that not in any cluster.
            3. The left sheet is the component distributions of each cluster, separately.

            The clustering algorithm is OPTICS, implemented by scikit-learn.
            https://scikit-learn.org/stable/modules/generated/sklearn.cluster.OPTICS.html

            Clustering algorithm details
                min_samples={1}
                min_cluster_size={2}
                xi={3}
                others=default

            """.format(QGRAIN_VERSION,
                       self.min_samples_input.value(),
                       self.min_cluster_size_input.value(),
                       self.xi_input.value())

        def write(row, col, value, style="normal_light"):
            cell = ws.cell(row+1, col+1, value=value)
            cell.style = style

        lines_of_desc = description.split("\n")
        for row, line in enumerate(lines_of_desc):
            write(row, 0, line, style="description")
        ws.column_dimensions[column_to_char(0)].width = 200

        ws = wb.create_sheet(self.tr("Typical Components"))
        write(0, 0, self.tr("Typical Component"), style="header")
        ws.column_dimensions[column_to_char(0)].width = 16
        for col, value in enumerate(classes_μm, 1):
            write(0, col, value, style="header")
            ws.column_dimensions[column_to_char(col)].width = 10
        for row, distribution in enumerate(typicals, 1):
            if row % 2 == 0:
                style = "normal_dark"
            else:
                style = "normal_light"
            write(row, 0, self.tr("Component{0}").format(row), style=style)
            for col, value in enumerate(distribution, 1):
                write(row, col, value, style=style)
            QCoreApplication.processEvents()

        for flag in flag_set:
            if flag == -1:
                ws = wb.create_sheet(self.tr("Not Clustered"), 2)
            else:
                ws = wb.create_sheet(self.tr("Cluster{0}").format(flag+1))

            write(0, 0, self.tr("Index"), style="header")
            ws.column_dimensions[column_to_char(0)].width = 16
            for col, value in enumerate(classes_μm, 1):
                write(0, col, value, style="header")
                ws.column_dimensions[column_to_char(col)].width = 10
            key = np.equal(flags, flag)
            for row, component in enumerate(self.stacked_components[key], 1):
                if row % 2 == 0:
                    style = "normal_dark"
                else:
                    style = "normal_light"
                write(row, 0, str(row), style=style)
                for col, value in enumerate(component, 1):
                    write(row, col, value, style=style)
                QCoreApplication.processEvents()

        wb.save(filename)
        wb.close()

    def on_save_clicked(self):
        if len(self.last_results) == 0:
            self.show_warning(self.tr("There is not an SSU result."))
            return

        filename, _ = self.file_dialog.getSaveFileName(
            self, self.tr("Choose a filename to save the typical components of SSU results"),
            None, f"{self.tr('Microsoft Excel')} (*.xlsx)")
        if filename is None or filename == "":
            return
        try:
            self.save_typical(filename)
            self.show_info(self.tr("The typical components have been saved."))
        except Exception as e:
            self.show_error(self.tr("Error raised while saving it to Excel file.\n    {0}").format(e.__str__()))
            return
