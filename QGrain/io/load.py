import csv
import logging
import os
from enum import Enum, unique

import numpy as np
import openpyxl
import xlrd
from typing import *

from ..models.dataset import Dataset, validate_classes, validate_distributions


CLASS_ROW_INDEX = 0
DISTRIBUTION_START_ROW_INDEX = 1
NAME_COL_INDEX = 0
SPACE_POSITION_COL_INDEXES = (1, 2, 3)

@unique
class ReadFileType(Enum):
    XLS = 0
    XLSX = 1
    CSV = 2


@unique
class ClassValueStyle(Enum):
    Midpoint = 0
    LeftBoundary = 1
    RightBoundary = 2
    EntireBoundary = 3


def get_file_type(filename: str):
    _, extension = os.path.splitext(filename)
    if extension == ".csv":
        return ReadFileType.CSV
    elif extension == ".xls":
        return ReadFileType.XLS
    elif extension == ".xlsx":
        return ReadFileType.XLSX
    else:
        raise NotImplementedError(extension)


def _get_raw_table(
        filename: str, file_type: ReadFileType, sheet_index: int,
        progress_callback: Callable[[float], None] = None,
        logger: logging.Logger = None):
    if logger is None:
        logger = logging.getLogger("QGrain")
    else:
        assert isinstance(logger, logging.Logger)
    logger.debug(f"Try to load the raw data table from [{file_type.name}] file: {filename}.")
    if file_type == ReadFileType.CSV:
        with open(filename, encoding="utf-8") as f:
            reader = csv.reader(f)
            raw_table = []
            for i, row_values in enumerate(reader):
                raw_table.append(row_values)
                if progress_callback is not None:
                    progress_callback(i / reader.line_num)

    elif file_type == ReadFileType.XLS:
        workbook: xlrd.Book = xlrd.open_workbook(filename)
        sheet = workbook.sheet_by_index(sheet_index)
        raw_table = []
        for row in range(sheet.nrows):
            raw_table.append(sheet.row_values(row))
            if progress_callback is not None:
                progress_callback(row / sheet.nrows)

    elif file_type == ReadFileType.XLSX:
        workbook: openpyxl.Workbook = openpyxl.load_workbook(filename, read_only=True, data_only=True)
        sheet = workbook[workbook.sheetnames[sheet_index]]
        raw_table = []
        for i, row_values in enumerate(sheet.values):
            raw_table.append(row_values)
            if progress_callback is not None:
                progress_callback(i / sheet.max_row)
    else:
        raise NotImplementedError(file_type)
    logger.debug("The raw data table has been loaded from this file.")
    return raw_table


def load_dataset(
        filename: str, dataset_name: str = None, sheet_index: int = 0,
        class_value_style: ClassValueStyle = ClassValueStyle.RightBoundary,
        contain_positions = False, skip_invalid_rows=True,
        progress_callback: Callable[[float], None] = None,
        logger: logging.Logger = None) -> Union[Dataset, None]:
    """
    Try to load the grain size dataset from a file.

    A data files of any types can be regarded as the table that contains rows and columns.

    For the file which contains grain size distributions, it should be the following format:
        * The first valid row should be the headers (i.e. the grain size classes in microns).
        * The following valid rows should be the distributions of samples corresponding to the grain size classes.
        * The first valid column should be the names of samples.
    To make it more flexible, you can use the settings to control the data loader.

    :param filename: The file path which contains the extension.
    :param dataset_name: The name of this dataset. If not indicate it, it will use the filename.
    :param sheet_index: If it is an Excel file (`*.xls` or `*.xlsx`), please indicate the sheet index,
        otherwise it will use the first sheet.
    :param class_value_style: It controls the behavior of processing the values in the class row.
    :param contain_positions: If `True`, it will use column 2 to 4 to represent the spatial position of the samples.
    :param skip_invalid_rows: If `True`, it will skip the invalid rows, rather than break up.
    :param progress_callback: The callable object which will be called at each step to report the progress.
    :param logger: The logger which will be used to log the information of loading.
    :return: If there is any exception raised, it will return `None`, else return a `Dataset` object.
    """
    assert isinstance(filename, str)
    assert len(filename) != 0
    if dataset_name is None:
        dataset_name = os.path.splitext(os.path.basename(filename))[0]
    else:
        assert isinstance(dataset_name, str)
        assert len(dataset_name) != 0
    assert isinstance(sheet_index, int)
    assert sheet_index >= 0
    assert isinstance(class_value_style, ClassValueStyle)
    assert isinstance(contain_positions, bool)
    assert isinstance(skip_invalid_rows, bool)
    assert isinstance(progress_callback, Callable) or progress_callback is None
    assert isinstance(logger, logging.Logger) or logger is None

    if logger is None:
        logger = logging.getLogger("QGrain")
    file_type = get_file_type(filename)
    if contain_positions:
        DISTRIBUTION_START_COL = max(SPACE_POSITION_COL_INDEXES) + 1
    else:
        DISTRIBUTION_START_COL = min(SPACE_POSITION_COL_INDEXES)
    try:
        raw_table = _get_raw_table(
            filename, file_type, sheet_index,
            progress_callback=None if progress_callback is None else lambda p: progress_callback(p * 0.5))
        class_values = raw_table[CLASS_ROW_INDEX][DISTRIBUTION_START_COL:]
        even_spaced = class_value_style != ClassValueStyle.EntireBoundary
        valid, array_or_msg = validate_classes(class_values, even_spaced=even_spaced)
        if valid:
            class_values = array_or_msg
            class_values_phi = -np.log2(class_values / 1000.0)
            average_class_interval_phi = np.mean(np.diff(class_values_phi))
            logger.debug(f"Grain size classes in microns: [{','.join([f'{x: 0.4f}' for x in class_values])}].")
        else:
            logger.error(f"The assigned series of grain size classes is invalid. {array_or_msg}")
            return None
        sample_names = []
        positions = []
        distributions = []
        for row, row_values in enumerate(raw_table[DISTRIBUTION_START_ROW_INDEX:], DISTRIBUTION_START_ROW_INDEX):
            # check if it's an empty row, i.e. the values all are empty string
            is_empty_row = True
            for frequency in row_values[DISTRIBUTION_START_COL:]:
                if frequency is not None and frequency != "":
                    is_empty_row = False
                    break
            # if it's an empty row, jump this row to process the next one
            if is_empty_row:
                logger.warning(f"Row {row} is empty, skip to next row.")
                continue

            sample_name = row_values[NAME_COL_INDEX]
            if sample_name is None:
                sample_name = "NONE"
                logger.warning(f"The sample name at row {row} is `None`, use 'NONE' instead.")
            elif not isinstance(sample_name, str):
                sample_name = str(sample_name)
                logger.warning(f"The sample name at row {row} is not `str`, use `str` to covert it.")
            elif len(sample_name) == 0:
                sample_name = "EMPTY"
                logger.warning(f"The sample name at row {row} is empty, use `EMPTY` to covert it.")

            if contain_positions:
                try:
                    position = [float(row_values[i]) for i in SPACE_POSITION_COL_INDEXES]
                except ValueError as e:
                    logger.error(f"The space position at row {row} contains invalid values, please check. {e}")
                if skip_invalid_rows:
                    continue
                else:
                    return None

            try:
                distribution = np.array(row_values[DISTRIBUTION_START_COL:], dtype=np.float32)
            except ValueError as e:
                logger.error(f"Can not convert the frequencies at row {row} to a numerical array, "
                             f"it may contains invalid values (e.g. text or empty cell). {e}")
                if skip_invalid_rows:
                    continue
                else:
                    return None
            sample_names.append(sample_name)
            if contain_positions:
                positions.append(position)
            distributions.append(distribution)
            # logger.debug(f"The validation of sample [{sample_name}] at row [{current_row}] is passed.")
            if progress_callback is not None:
                progress = row / len(raw_table) * 0.5 + 0.5
                progress_callback(progress)

        valid, array_or_msg = validate_distributions(distributions)
        if not valid:
            logger.error(f"There is at least one grain size distribution is invalid. {array_or_msg}")
            return None
        distributions = array_or_msg

        boundaries = np.zeros(len(class_values) + 1)
        if class_value_style == ClassValueStyle.Midpoint:
            boundaries[:-1] = class_values_phi - average_class_interval_phi / 2
            boundaries[-1] = boundaries[-2] + average_class_interval_phi
        elif class_value_style == ClassValueStyle.LeftBoundary:
            boundaries[:-1] = class_values_phi
            boundaries[-1] = boundaries[-2] + average_class_interval_phi
        elif class_value_style == ClassValueStyle.RightBoundary:
            boundaries[1:] = class_values_phi
            boundaries[0] = boundaries[1] - average_class_interval_phi
        elif class_value_style == ClassValueStyle.EntireBoundary:
            boundaries = class_values_phi
        else:
            raise NotImplementedError(class_value_style)
        assert len(boundaries) == distributions.shape[1]

        dataset = Dataset(dataset_name, sample_names, boundaries, distributions,
                          positions=positions if contain_positions else None)
        if progress_callback is not None:
            progress_callback(1.0)
        logger.info("This dataset has been loaded successfully.")
        return dataset

    except IOError as e:
        logger.error(f"Can not open this file. {e}")
        return None
    except IndexError as e:
        logger.error(f"The row or column index is out of range, please check. {e}")
        return None
