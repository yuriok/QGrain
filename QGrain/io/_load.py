import csv
import logging
import os
import typing
from enum import Enum, unique

import numpy as np
import openpyxl
import xlrd

from ..models import GrainSizeDataset


class DataLayoutError(Exception):
    """Raises while the data layout settings are invalid."""
    pass


class DataLayoutSetting:
    """
    The class to represent the layout setting of raw data file.
    All types of raw data files can be regarded as the table(s) that contains rows and columns.
    For the file which contains grain size distributions, it should be the following format:
        * The first valid row should be the headers (i.e. the grain size classes).
        * The following valid rows should be the distributions of samples corresponding to the grain size classes.
        * The first valid column shoud be the names of samples.
    To make it more flexible, we use this setting to control the data loader.
    """
    def __init__(self, classes_row=0, sample_names_column=0,
                 distribution_start_row=1, distribution_start_column=1):
        # make sure the types are int in other codes
        assert isinstance(classes_row, int)
        assert isinstance(sample_names_column, int)
        assert isinstance(distribution_start_row, int)
        assert isinstance(distribution_start_column, int)
        # handle these errors at front-end
        if classes_row < 0 or \
            sample_names_column < 0 or \
            distribution_start_row < 0 or \
            distribution_start_column < 0:
            raise DataLayoutError("Index must be non-negative.")
        if classes_row >= distribution_start_row:
            raise DataLayoutError("The start row index of distribution must be greater than the row index of classes.")
        if sample_names_column >= distribution_start_column:
            raise DataLayoutError("The start column index of distribution must be greater than the column index of sample names.")
        self.__classes_row = classes_row
        self.__sample_names_column = sample_names_column
        self.__distribution_start_row = distribution_start_row
        self.__distribution_start_column = distribution_start_column

    @property
    def classes_row(self) -> int:
        return self.__classes_row

    @property
    def sample_names_column(self) -> int:
        return self.__sample_names_column

    @property
    def distribution_start_row(self) -> int:
        return self.__distribution_start_row

    @property
    def distribution_start_column(self) -> int:
        return self.__distribution_start_column


@unique
class ReadFileType(Enum):
    XLS = 0
    XLSX = 1
    CSV = 2


def get_type_by_name(filename: str):
    _, extension = os.path.splitext(filename)
    if extension == ".csv":
        return ReadFileType.CSV
    elif extension == ".xls":
        return ReadFileType.XLS
    elif extension == ".xlsx":
        return ReadFileType.XLSX
    else:
        raise NotImplementedError(extension)


def _get_raw_table(
    filename: str, file_type: ReadFileType, sheet_index: int,
    progress_callback: typing.Callable = None, logger: logging.Logger = None):
    if logger is None:
        logger = logging.getLogger("QGrain")
    else:
        assert isinstance(logger, logging.Logger)
    logger.debug(f"Try to load raw data table from [{file_type.name}] file: {filename}.")
    if file_type == ReadFileType.CSV:
        with open(filename, encoding="utf-8") as f:
            reader = csv.reader(f)
            raw_table = []
            for i, row_values in enumerate(reader):
                raw_table.append(row_values)
                if progress_callback is not None:
                    progress_callback(i / reader.line_num)

    elif file_type == ReadFileType.XLS:
        workbook = xlrd.open_workbook(filename) # type: xlrd.Book
        sheet = workbook.sheet_by_index(sheet_index)
        raw_table = []
        for row in range(sheet.nrows):
            raw_table.append(sheet.row_values(row))
            if progress_callback is not None:
                progress_callback(row / sheet.nrows)

    elif file_type == ReadFileType.XLSX:
        workbook = openpyxl.load_workbook(filename, read_only=True, data_only=True) # type: openpyxl.Workbook
        sheet = workbook[workbook.sheetnames[sheet_index]]
        raw_table = []
        for i, row_values in enumerate(sheet.values):
            raw_table.append(row_values)
            if progress_callback is not None:
                progress_callback(i / sheet.max_row)
    else:
        raise NotImplementedError(file_type)
    logger.debug("Raw data table has been loaded from this file.")
    return raw_table


def load_dataset(filename: str,
                 sheet_index: int = 0,
                 layout: DataLayoutSetting = None,
                 progress_callback: typing.Callable = None,
                 logger: logging.Logger = None):
    """
    Try to load the grain size dataset from the file.

    ## Parameters

    filename: `str`
        The file path which contains the extension.
    sheet_index: `int`
        If it is an Excel file (`*.xls` or `*.xlsx`), please indicate the sheet index, otherwise it will use the first sheet.
    layout: `DataLayoutSetting`
        The layout of the grain size data stored in the sheet.
    logger: `logging.Logger`
        The logger which will be used to log the information of loading.

    ## Returns

    result: `None` or `GrainSizeDataset`
        If there is any exception raised, it will return `None`.

    """
    assert isinstance(filename, str)
    assert len(filename) != 0
    assert isinstance(sheet_index, int)
    assert sheet_index >= 0
    if layout is None:
        layout = DataLayoutSetting()
    else:
        assert isinstance(layout, DataLayoutSetting)
    if logger is None:
        logger = logging.getLogger("QGrain")
    else:
        assert isinstance(logger, logging.Logger)

    file_type = get_type_by_name(filename)
    if progress_callback is not None:
        _callback = lambda progress: progress_callback(progress*0.5)
    else:
        _callback = None
    raw_table = _get_raw_table(filename, file_type, sheet_index, progress_callback=_callback)
    try:
        class_values = raw_table[layout.classes_row][layout.distribution_start_column:]
        classes_μm = np.array(class_values, dtype=np.float64)
    except IndexError as e:
        logger.exception(f"The row index of classes [{layout.classes_row}] or start column index [{layout.distribution_start_column}] of distribution is out of range, please check.", stack_info=True)
        return
    except ValueError as e:
        logger.exception(f"Can not convert the values of classes to a numerical array, it may contains invalid values (e.g. text).", stack_info=True)
        return
    logger.debug(f"Grain size classes in μm: [{','.join([f'{x: 0.4f}' for x in classes_μm])}].")
    try:
        GrainSizeDataset.validate_classes_μm(classes_μm)
    except Exception as e:
        logger.exception(f"The array of grain size classes is invalid, please check.", stack_info=True)
        return
    logger.debug("The array of grain size classes is valid.")

    try:
        names = []
        distributions = []
        for current_row, row_values in enumerate(raw_table[layout.distribution_start_row:], layout.distribution_start_row+1):
            # check if it's a empty row, i.e. the values all are empty string
            is_empty_row = True
            for distribution_value in row_values[layout.distribution_start_column:]:
                if distribution_value != "" and distribution_value is not None:
                    is_empty_row = False
                    break
            # if it's a empty row, jump this row to process the next one
            if is_empty_row:
                logger.warning(f"Row [{current_row}] is empty, skip to next row.")
                continue

            sample_name = row_values[layout.sample_names_column]
            # logger.debug(f"Processing the row [{current_row}].")
            if sample_name is None:
                sample_name = "NONE"
                logger.warning(f"The sample name is invalid, use 'NONE' instead.")
            # users may use pure number as the sample name
            elif not isinstance(sample_name, str):
                sample_name = str(sample_name)
                logger.warning(f"The sample name is not text (may be a number), convert it to text.")
            elif sample_name == "":
                sample_name = "EMPTY"
                logger.warning(f"The sample name is a empty text, use 'EMPTY' instead.")
            try:
                distribution = np.array(row_values[layout.distribution_start_column:], dtype=np.float64)
            except ValueError as e:
                logger.exception(f"Can not convert the distribution values at row [{current_row}] to a numerical array, it may contains invalid values (e.g. text or empty cell).", stack_info=True)
                return
            try:
                GrainSizeDataset.validate_distribution(distribution)
            except Exception as e:
                logger.exception(f"The distribution array of sample [{sample_name}] at row [{current_row}] is invalid.", stack_info=True)
                return
            names.append(sample_name)
            distributions.append(distribution)
            # logger.debug(f"The validation of sample [{sample_name}] at row [{current_row}] is passed.")
            if progress_callback is not None:
                progress = current_row / len(raw_table) * 0.5 + 0.5
                progress_callback(progress)
    except IndexError as e:
        logger.exception("At least one row or column index of data layout is out of range, please check.", stack_info=True)
        return
    try:
        dataset = GrainSizeDataset()
        dataset.add_batch(classes_μm, names, distributions, need_validation=False)
    except Exception as e:
        logger.exception("An unknown exception was raised. Please check the logs for more details.", stack_info=True)
        return
    logger.info("This dataset has been loaded successfully.")
    return dataset
