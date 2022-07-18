__all__ = [
    "column_to_char",
    "to_cell_name",
    "prepare_styles"]

import openpyxl
from openpyxl.styles import (Alignment, Font, NamedStyle, PatternFill)


def column_to_char(column_index: int):
    column = column_index + 1
    column_str = str()
    while column != 0:
        res = column % 26
        if res == 0:
            res = 26
            column -= 26
        column_str = chr(ord('A') + res - 1) + column_str
        column = column // 26
    return column_str

def to_cell_name(row: int, column: int):
    return "{0}{1}".format(column_to_char(column), row+1)

def prepare_styles(wb: openpyxl.Workbook):
    normal_light_style = NamedStyle(name="normal_light")
    normal_light_style.font = Font(size=12, name="Arial", color="000000")
    normal_light_style.alignment = Alignment(
        horizontal='center',
        vertical='center',
        text_rotation=0,
        wrap_text=False,
        shrink_to_fit=False,
        indent=0)
    normal_light_style.fill = PatternFill(
        patternType="solid",
        fgColor="c4cbcf")
    wb.add_named_style(normal_light_style)

    normal_dark_style = NamedStyle(name="normal_dark")
    normal_dark_style.font = Font(size=12, name="Arial", color="000000")
    normal_dark_style.alignment = Alignment(
        horizontal='center',
        vertical='center',
        text_rotation=0,
        wrap_text=False,
        shrink_to_fit=False,
        indent=0)
    normal_dark_style.fill = PatternFill(
        patternType="solid",
        fgColor="b2bbbe")
    wb.add_named_style(normal_dark_style)

    warning_style = NamedStyle(name="warning")
    warning_style.font = Font(size=12, name="Arial", color="ffffff")
    warning_style.alignment = Alignment(
        horizontal='center',
        vertical='center',
        text_rotation=0,
        wrap_text=False,
        shrink_to_fit=False,
        indent=0)
    warning_style.fill = PatternFill(
        patternType="solid",
        fgColor="ff9900")
    wb.add_named_style(warning_style)

    header_style = NamedStyle(name="header")
    header_style.font = Font(size=14, bold=True, name="Times New Roman", color="ffffff")
    header_style.alignment = Alignment(
        horizontal='center',
        vertical='center',
        text_rotation=0,
        wrap_text=True,
        shrink_to_fit=False,
        indent=0)
    header_style.fill = PatternFill(
        patternType="solid",
        fgColor="475164")
    wb.add_named_style(header_style)

    description_style = NamedStyle(name="description")
    description_style.font = Font(size=14, bold=True, name="Times New Roman", color="ffffff")
    description_style.alignment = Alignment(
        horizontal='left',
        vertical='center',
        text_rotation=0,
        wrap_text=False,
        shrink_to_fit=False,
        indent=0)
    description_style.fill = PatternFill(
        patternType="solid",
        fgColor="475164")
    wb.add_named_style(description_style)
